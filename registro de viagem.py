"""
Sistema de Registro de Viagens
Obrigatorias : pip install pyperclip openpyxl cryptography
Opcionais    : pip install sv-ttk reportlab
"""
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from datetime import datetime
import json, os, re, shutil, uuid, webbrowser, hashlib, html, threading
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from pathlib import Path
import pyperclip
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

try:
    import sv_ttk          # type: ignore[import]
    _SV_TTK = True
except ImportError:
    _SV_TTK = False

try:
    from reportlab.lib.pagesizes import A4            # type: ignore[import]
    from reportlab.pdfgen import canvas as rl_canvas  # type: ignore[import]
    _REPORTLAB = True
except ImportError:
    _REPORTLAB = False

try:
    from cryptography.fernet import Fernet            # type: ignore[import]
    import base64
    _CRYPTO = True
except ImportError:
    _CRYPTO = False

# ── Constantes ────────────────────────────────────────────────────────────────
_BASE_DIR   = Path(__file__).parent
DATA_FILE   = str(_BASE_DIR / "viagens.json")
CONFIG_FILE = str(_BASE_DIR / "config.json")
BACKUP_DIR  = str(_BASE_DIR / "backups")
ATTACH_DIR  = str(_BASE_DIR / "anexos")
BACKUP_MAX  = 30  # quantos backups manter

DEFAULT_CONFIG = {
    "valor_diaria":       150.00,
    "backup_automatico":  True,
    "tema":               "light",
    "email_remetente":    "",
    "email_senha":        "",
    "email_smtp":         "smtp.gmail.com",
    "email_porta":        587,
    "email_destinatario": "",
    "auto_email":         False,
    "cfg_hash":           "",
    "cfg_salt":           "",
    "cfg_user":           "",
}

MESES = ["Todos","Janeiro","Fevereiro","Marco","Abril","Maio","Junho",
         "Julho","Agosto","Setembro","Outubro","Novembro","Dezembro"]
TIPOS_PASSAGEM   = ["Aereo","Rodoviario","Ferroviario","Maritimo","Outro"]
TIPOS_TRANSPORTE = ["Uber","99","Taxi","Onibus","Metro","VLT","Trem","Outro"]

# ── Helpers de segurança ──────────────────────────────────────────────────────
def _hash_senha(txt: str, salt: bytes | None = None) -> tuple[str, str]:
    """Retorna (hash_hex, salt_hex) usando scrypt com salt aleatório."""
    if salt is None:
        salt = os.urandom(16)
    h = hashlib.scrypt(txt.strip().encode("utf-8"), salt=salt, n=2**14, r=8, p=1)
    return h.hex(), salt.hex()

def _verificar_senha(txt: str, hash_hex: str, salt_hex: str) -> bool:
    h, _ = _hash_senha(txt, bytes.fromhex(salt_hex))
    return h == hash_hex

def _fernet_key(senha: str, salt_hex: str) -> bytes:
    """Deriva uma chave Fernet a partir da senha do app."""
    salt = bytes.fromhex(salt_hex)
    k = hashlib.scrypt(senha.strip().encode("utf-8"), salt=salt, n=2**14, r=8, p=1, dklen=32)
    return base64.urlsafe_b64encode(k)

def _cifrar(texto: str, senha: str, salt_hex: str) -> str:
    if not _CRYPTO or not texto:
        return texto
    f = Fernet(_fernet_key(senha, salt_hex))
    return f.encrypt(texto.encode("utf-8")).decode("ascii")

def _decifrar(token: str, senha: str, salt_hex: str) -> str:
    if not _CRYPTO or not token:
        return token
    try:
        f = Fernet(_fernet_key(senha, salt_hex))
        return f.decrypt(token.encode("ascii")).decode("utf-8")
    except Exception:
        return ""

def _abrir_arquivo(path: str) -> None:
    """Abre um arquivo com o aplicativo padrão do sistema."""
    if os.name == "nt":
        os.startfile(path)
    elif os.name == "posix":
        import subprocess
        opener = "open" if __import__("sys").platform == "darwin" else "xdg-open"
        subprocess.Popen([opener, path])

# ── Paleta — 3 temas ──────────────────────────────────────────────────────────
_PALETAS = {
    "light": {
        "bg":             "#E8ECF0",   # cinza levemente azulado (harmonia com logo)
        "surface":        "#F3F5F7",
        "surface2":       "#DDE2E8",
        "border":         "#B8C2CC",
        "accent":         "#4A7A9B",   # azul da marca MF
        "success":        "#2E7D32",
        "error":          "#C62828",
        "text":           "#1A2530",
        "text_dim":       "#5A6978",
        "header":         "#4A7A9B",   # header na cor da marca
        "btn":            "#3A6282",   # azul mais escuro para botões
        "btn_hover":      "#4A7A9B",
        "btn_ok":         "#3A6282",
        "btn_ok_h":       "#4A7A9B",
        "btn_danger":     "#3A6282",
        "btn_danger_h":   "#4A7A9B",
        "btn_neutral":    "#5A7A8A",
        "btn_neutral_h":  "#6A8A9A",
        "btn_tx_default": "#E8F4FF",   # branco azulado
        "btn_tx_ok":      "#B9F6CA",   # verde água
        "btn_tx_danger":  "#FF8A80",   # vermelho suave
        "btn_tx_neutral": "#CFD8DC",   # cinza azulado
        "titlebar":       0x009B7A4A,  # COLORREF: #4A7A9B em BGR
    },
    "dark": {
        "bg":             "#2B2B2B",
        "surface":        "#3C3C3C",
        "surface2":       "#333333",
        "border":         "#606060",
        "accent":         "#E0E0E0",
        "success":        "#6FCF97",
        "error":          "#EB5757",
        "text":           "#F0F0F0",
        "text_dim":       "#AAAAAA",
        "header":         "#1E1E1E",
        "btn":            "#505050",
        "btn_hover":      "#686868",
        "btn_ok":         "#505050",
        "btn_ok_h":       "#686868",
        "btn_danger":     "#505050",
        "btn_danger_h":   "#686868",
        "btn_neutral":    "#3C3C3C",
        "btn_neutral_h":  "#555555",
        "btn_tx_default": "#90CAF9",
        "btn_tx_ok":      "#A5D6A7",
        "btn_tx_danger":  "#EF9A9A",
        "btn_tx_neutral": "#B0BEC5",
        "titlebar":       0x001E1E1E,
    },
}

# C aponta para a paleta do tema ativo — reatribuído em _apply_theme
C = dict(_PALETAS["dark"])

ICONS = {
    "add":     "⊕",   # ⊕
    "remove":  "⊖",   # ⊖
    "edit":    "✎",   # ✎
    "save":    "✔",   # ✔
    "cancel":  "✕",   # ✕
    "copy":    "⎘",   # ⎘
    "email":   "✉",   # ✉
    "excel":   "⊞",   # ⊞
    "backup":  "⇩",   # ⇩
    "refresh": "↺",   # ↺
    "open":    "↗",   # ↗
    "flight":  "✈",   # ✈
    "gear":    "⚙",   # ⚙
    "lock":    "⊟",   # ⊟
    "key":     "❖",   # ❖
    "pix":     "◈",   # ◈
    "doc":     "▤",   # ▤
    "clear":   "⌫",   # ⌫
    "attach":  "⚬",   # ⚬
    "hotel":   "⌂",   # ⌂
    "car":     "▶",   # ▶
    "dark":    "◑",   # ◑
    "black":   "●",   # ●
    "light":   "○",   # ○
    "link":    "⚭",   # ⚭  (repurposed)
    "nf":      "▦",   # ▦
}

_FONT      = "Segoe UI"
_FONT_MONO = "Consolas"

TEMA_CICLO  = ["light", "dark"]
TEMA_LABELS = {
    "light": f"{ICONS['light']}  Claro",
    "dark":  f"{ICONS['dark']}   Escuro",
}

_TEMA_ATUAL = "dark"   # atualizado por _apply_theme antes de criar widgets

def _lighten(hex_color: str, factor: float = 0.35) -> str:
    """Clareia uma cor hex para o destaque do topo das barras."""
    hex_color = hex_color.lstrip("#")
    r, g, b = int(hex_color[0:2], 16), int(hex_color[2:4], 16), int(hex_color[4:6], 16)
    r = min(255, int(r + (255 - r) * factor))
    g = min(255, int(g + (255 - g) * factor))
    b = min(255, int(b + (255 - b) * factor))
    return f"#{r:02X}{g:02X}{b:02X}"

def _btn_colors(danger=False, ok=False, neutral=False, **_):
    bg  = C["btn_danger"]  if danger  else C["btn_ok"]  if ok  else C["btn_neutral"] if neutral else C["btn"]
    hv  = C["btn_danger_h"]if danger  else C["btn_ok_h"]if ok  else C["btn_neutral_h"]if neutral else C["btn_hover"]
    tx  = C.get("btn_tx_danger" if danger else "btn_tx_ok" if ok else "btn_tx_neutral" if neutral else "btn_tx_default",
                C["text"])
    return bg, hv, tx


class ModernButton(tk.Button):
    """Botão tk com hover animado e texto vibrante."""
    def __init__(self, master, text="", icon="", bg=None, hover=None,
                 fg=None, danger=False, ok=False, neutral=False, **kw):
        _bg, _hv, _fg = _btn_colors(danger, ok, neutral)
        _bg = bg    or _bg
        _hv = hover or _hv
        _fg = fg    or _fg
        label = f"{icon}  {text}" if icon else text
        super().__init__(
            master, text=label, bg=_bg, fg=_fg, relief="flat",
            activebackground=_hv, activeforeground=_fg,
            cursor="hand2", padx=10, pady=5,
            font=(_FONT, 9, "bold"), bd=0, highlightthickness=0, **kw)
        self._bg = _bg
        self._hv = _hv
        self.bind("<Enter>", lambda e: self.config(bg=self._hv))
        self.bind("<Leave>", lambda e: self.config(bg=self._bg))


# ─────────────────────────────────────────────────────────────────────────────
class AppViagens(tk.Tk):
    def __init__(self):
        super().__init__()
        self.cfg          = self._load_cfg()
        self._passagens   = []
        self._transportes = []
        self._materiais   = []
        self._cfg_unlocked = False
        self._cfg_senha    = ""   # senha em memória para cifrar/decifrar email_senha
        self.title("Registro de Viagens")
        self.resizable(True, True)
        self.state("zoomed")
        self._apply_theme(self.cfg["tema"])
        self._backup_auto()
        self._build_ui()

    # ── Config ────────────────────────────────────────────────────────────────
    def _load_cfg(self):
        if os.path.exists(CONFIG_FILE):
            try:
                with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                    c = json.load(f)
                for k, v in DEFAULT_CONFIG.items():
                    c.setdefault(k, v)
                return c
            except (json.JSONDecodeError, IOError):
                pass
        return DEFAULT_CONFIG.copy()

    def _save_cfg(self):
        try:
            with open(CONFIG_FILE, "w", encoding="utf-8") as f:
                json.dump(self.cfg, f, ensure_ascii=False, indent=4)
        except IOError as e:
            messagebox.showerror("Erro", f"Nao foi possivel salvar config:\n{e}")

    # ── Título do Windows (DWM) ───────────────────────────────────────────────
    def _apply_titlebar(self):
        try:
            import ctypes
            hwnd = ctypes.windll.user32.GetParent(self.winfo_id())
            if not hwnd:
                hwnd = self.winfo_id()
            dark_flag = 0 if self._tema_atual == "light" else 1
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, 20, ctypes.byref(ctypes.c_int(dark_flag)), 4)
            color = C.get("titlebar", 0x00D0D0D0)
            ctypes.windll.dwmapi.DwmSetWindowAttribute(
                hwnd, 35, ctypes.byref(ctypes.c_int(color)), 4)
        except Exception:
            pass

    # ── Tema ──────────────────────────────────────────────────────────────────
    def _apply_theme(self, tema):
        global C, _TEMA_ATUAL
        if tema not in _PALETAS:
            tema = "dark"
        self._tema_atual = tema
        _TEMA_ATUAL = tema
        C = dict(_PALETAS[tema])

        s  = ttk.Style(self)
        av = s.theme_names()
        base = "clam" if "clam" in av else av[0]
        s.theme_use(base)

        bg, surf, brd = C["bg"], C["surface"], C["border"]
        tx, td        = C["text"], C["text_dim"]
        acc           = C["accent"]
        hdr           = C["header"]

        s.configure(".",
            background=bg, foreground=tx,
            fieldbackground=surf, insertcolor=acc,
            troughcolor=surf, bordercolor=brd,
            darkcolor=surf, lightcolor=surf,
            font=(_FONT, 9))
        for w in ("TLabel", "TFrame", "TLabelframe", "TLabelframe.Label"):
            s.configure(w, background=bg, foreground=tx, font=(_FONT, 9))
        s.configure("TLabelframe", bordercolor=brd, relief="groove")
        s.configure("TLabelframe.Label",
            background=bg, foreground=acc, font=(_FONT, 9, "bold"))
        s.configure("TNotebook", background=bg, borderwidth=0)
        s.configure("TNotebook.Tab",
            background=surf, foreground=td,
            padding=[14, 6], font=(_FONT, 9))
        s.map("TNotebook.Tab",
            background=[("selected", bg),   ("active", brd)],
            foreground=[("selected", acc),  ("active", tx)])
        s.configure("TEntry",
            fieldbackground=surf, foreground=tx,
            insertcolor=acc, bordercolor=brd,
            font=(_FONT, 9), relief="flat", padding=4)
        s.configure("TCombobox",
            fieldbackground=surf, foreground=tx,
            background=surf, arrowcolor=acc,
            bordercolor=brd, font=(_FONT, 9))
        s.map("TCombobox",
            fieldbackground=[("readonly", surf)],
            foreground=[("readonly", tx)])
        s.configure("TScrollbar",
            background=surf, troughcolor=bg,
            arrowcolor=td, bordercolor=bg)
        s.configure("Treeview",
            background=surf, foreground=tx,
            fieldbackground=surf, font=(_FONT, 9), rowheight=24)
        s.configure("Treeview.Heading",
            background=hdr, foreground=acc,
            font=(_FONT, 9, "bold"), relief="flat")
        s.map("Treeview",
            background=[("selected", brd)],
            foreground=[("selected", tx)])
        s.configure("TCheckbutton", background=bg, foreground=tx, font=(_FONT, 9))
        s.configure("TRadiobutton", background=bg, foreground=tx, font=(_FONT, 9))
        self.configure(bg=bg)
        self.after(100, self._apply_titlebar)

    # ── Backup ────────────────────────────────────────────────────────────────
    def _backup_auto(self):
        if not self.cfg.get("backup_automatico") or not os.path.exists(DATA_FILE):
            return
        os.makedirs(BACKUP_DIR, exist_ok=True)
        dest = os.path.join(BACKUP_DIR, f"backup_{datetime.now().strftime('%Y%m%d')}.json")
        if not os.path.exists(dest):
            try: shutil.copy2(DATA_FILE, dest)
            except IOError: pass
        self._limpar_backups_antigos()

    def _limpar_backups_antigos(self):
        try:
            arquivos = sorted(Path(BACKUP_DIR).glob("backup_*.json"),
                              key=lambda p: p.stat().st_mtime, reverse=True)
            for arq in arquivos[BACKUP_MAX:]:
                arq.unlink(missing_ok=True)
        except OSError:
            pass

    def _backup_manual(self):
        if not os.path.exists(DATA_FILE):
            messagebox.showwarning("Aviso", "Nenhum arquivo de dados.")
            return
        os.makedirs(BACKUP_DIR, exist_ok=True)
        ts   = datetime.now().strftime("%Y%m%d_%H%M%S")
        dest = os.path.join(BACKUP_DIR, f"backup_{ts}.json")
        try:
            shutil.copy2(DATA_FILE, dest)
            messagebox.showinfo("Backup", f"Backup criado:\n{os.path.abspath(dest)}")
        except IOError as e:
            messagebox.showerror("Erro", str(e))

    # ── Anexos ────────────────────────────────────────────────────────────────
    def _salvar_anexo(self, reg_id: str, filepath: str) -> str:
        dest_dir = Path(ATTACH_DIR) / reg_id
        dest_dir.mkdir(parents=True, exist_ok=True)
        dest = dest_dir / Path(filepath).name
        shutil.copy2(filepath, dest)
        return str(dest)

    # ── UI Principal ──────────────────────────────────────────────────────────
    def _build_ui(self):
        hdr = tk.Frame(self, bg=C["header"], height=64)
        hdr.pack(fill=tk.X)
        hdr.pack_propagate(False)

        # Slot da logo (esquerda)
        self.lbl_logo = tk.Label(hdr, bg=C["header"], width=52, height=52,
                                 text="MF", font=(_FONT, 18, "bold"), fg="#FFFFFF")
        self.lbl_logo.place(x=10, rely=0.5, anchor="w")

        # Título centralizado
        tk.Label(hdr, text="MF VIAGENS E HOTÉIS",
                 font=(_FONT, 17, "bold"),
                 bg=C["header"], fg="#FFFFFF").place(relx=0.5, rely=0.5, anchor="center")

        # Botão de tema (direita)
        prox = TEMA_CICLO[(TEMA_CICLO.index(self.cfg["tema"]) + 1) % len(TEMA_CICLO)]
        self.btn_tema = ModernButton(
            hdr, text=TEMA_LABELS[prox], neutral=True,
            command=self._toggle_tema)
        self.btn_tema.place(relx=1.0, rely=0.5, anchor="e", x=-12)

        # Notebook
        self.nb = ttk.Notebook(self)
        self.nb.pack(fill=tk.BOTH, expand=True, padx=6, pady=(4, 0))
        self.nb.bind("<<NotebookTabChanged>>", self._on_tab)

        self.tab_reg    = ttk.Frame(self.nb)
        self.tab_hist   = ttk.Frame(self.nb, padding=10)
        self.tab_painel = ttk.Frame(self.nb, padding=10)
        self.tab_cfg    = ttk.Frame(self.nb, padding=10)
        self.nb.add(self.tab_reg,    text=f"  {ICONS['doc']}  Novo Registro  ")
        self.nb.add(self.tab_hist,   text=f"  {ICONS['refresh']}  Historico  ")
        self.nb.add(self.tab_painel, text=f"  {ICONS['excel']}  Painel  ")
        self.nb.add(self.tab_cfg,    text=f"  {ICONS['gear']}  Configuracoes  ")

        self._build_registro()
        self._build_historico()
        self._build_painel()
        self._build_cfg_lock()
        self.after(200, self._set_logo)

        # Status bar
        sb = tk.Frame(self, bg=C["header"], height=28)
        sb.pack(fill=tk.X, side=tk.BOTTOM)
        sb.pack_propagate(False)
        self.status_var = tk.StringVar(value="Pronto.")
        tk.Label(sb, textvariable=self.status_var,
                 font=(_FONT, 9), bg=C["header"], fg="#DDEEFF").pack(
            side=tk.LEFT, padx=12)
        self._refresh_status()

    def _set_logo(self, path: str = ""):
        """Carrega logo no header. Busca arquivos de imagem na pasta do projeto."""
        if not path:
            # busca por qualquer arquivo de logo na pasta
            candidates = []
            for pattern in ("logo*.*", "LOGO*.*", "Logo*.*"):
                candidates.extend(_BASE_DIR.glob(pattern))
            candidates = [p for p in candidates
                          if p.suffix.lower() in (".png",".jpg",".jpeg",".gif",".bmp")]
            if not candidates:
                return
            path = str(candidates[0])
        try:
            from PIL import Image, ImageTk  # type: ignore[import]
            img = Image.open(path).resize((54, 54), Image.LANCZOS)
            self._logo_img = ImageTk.PhotoImage(img)
            self.lbl_logo.config(image=self._logo_img, text="", bg=C["header"])
        except ImportError:
            # Pillow não instalado — tenta PhotoImage nativo (só PNG/GIF)
            if path.lower().endswith((".png", ".gif")):
                try:
                    raw = tk.PhotoImage(file=path)
                    factor = max(1, raw.width() // 54)
                    self._logo_img = raw.subsample(factor, factor)
                    self.lbl_logo.config(image=self._logo_img, text="", bg=C["header"])
                except Exception:
                    pass
        except Exception:
            pass

    def _toggle_tema(self):
        idx  = TEMA_CICLO.index(self.cfg.get("tema", "dark"))
        novo = TEMA_CICLO[(idx + 1) % len(TEMA_CICLO)]
        self.cfg["tema"] = novo
        self._save_cfg()
        self._apply_theme(novo)
        prox = TEMA_CICLO[(TEMA_CICLO.index(novo) + 1) % len(TEMA_CICLO)]
        self.btn_tema.config(text=TEMA_LABELS[prox],
                             bg=C["btn_neutral"], activebackground=C["btn_neutral_h"])
        if hasattr(self, "var_tema_cfg"):
            self.var_tema_cfg.set(novo)

    # ── Tab Registro (scroll) ─────────────────────────────────────────────────
    def _build_registro(self):
        outer  = self.tab_reg
        canvas = tk.Canvas(outer, highlightthickness=0, bg=C["bg"])
        vsb    = ttk.Scrollbar(outer, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        inner   = ttk.Frame(canvas, padding=(14, 10))
        self._reg_win = canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: canvas.configure(
            scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(
            self._reg_win, width=e.width - 4))
        def _scroll(e):
            canvas.yview_scroll(int(-1*(e.delta/120)), "units")
        canvas.bind("<Enter>",  lambda e: canvas.bind_all("<MouseWheel>", _scroll))
        canvas.bind("<Leave>",  lambda e: canvas.unbind_all("<MouseWheel>"))
        inner.columnconfigure(1, weight=1)

        row = 0
        row = self._sec_header(inner, row, "Dados da Viagem", "A")
        row = self._sec_basico(inner, row)
        row = self._sec_header(inner, row, "Hospedagem", "B")
        row = self._sec_hospedagem(inner, row)
        row = self._sec_header(inner, row, "Passagens", "C")
        row = self._sec_passagens(inner, row)
        row = self._sec_header(inner, row, "Transporte Local / Uber", "D")
        row = self._sec_transportes(inner, row)
        row = self._sec_header(inner, row, "Materiais e Despesas", "E")
        row = self._sec_materiais(inner, row)
        row = self._sec_header(inner, row, "Resumo de Custos", "F")
        row = self._sec_resumo(inner, row)
        row = self._sec_header(inner, row, "Links Cadastrados", "G")
        row = self._sec_links(inner, row)
        row = self._sec_header(inner, row, "PIX e Observacoes", "H")
        row = self._sec_pix_obs(inner, row)
        self._sec_botoes(inner, row)

    def _sec_header(self, f, row, titulo, letra):
        fr = tk.Frame(f, bg=C["bg"])
        fr.grid(row=row, column=0, columnspan=5, sticky="ew", pady=(14, 3))
        tk.Label(fr, text=f" {letra} ", font=(_FONT, 8, "bold"),
                 bg=C["accent"], fg=C["header"], padx=6, pady=3).pack(side=tk.LEFT)
        tk.Label(fr, text=f"  {titulo}",
                 font=(_FONT, 10, "bold"),
                 bg=C["bg"], fg=C["accent"]).pack(side=tk.LEFT)
        tk.Frame(fr, bg=C["border"], height=1).pack(
            side=tk.LEFT, fill=tk.X, expand=True, padx=(10, 0), pady=6)
        return row + 1

    # ── Secao A: Dados basicos ────────────────────────────────────────────────
    def _sec_basico(self, f, row):
        def lbl(txt): return ttk.Label(f, text=txt, font=(_FONT, 9, "bold"))

        lbl("Nome do Viajante:").grid(row=row, column=0, sticky="w", pady=4)
        self.e_nome = ttk.Entry(f, width=50)
        self.e_nome.grid(row=row, column=1, columnspan=4, sticky="w", pady=4)
        row += 1

        lbl("Destino:").grid(row=row, column=0, sticky="w", pady=4)
        self.e_destino = ttk.Entry(f, width=50)
        self.e_destino.grid(row=row, column=1, columnspan=4, sticky="w", pady=4)
        row += 1

        # Partida — data e hora na mesma linha sem label intermediário
        lbl("Partida:").grid(row=row, column=0, sticky="w", pady=4)
        fr_ida = ttk.Frame(f)
        fr_ida.grid(row=row, column=1, columnspan=4, sticky="w", pady=4)
        self.e_ida = ttk.Entry(fr_ida, width=13)
        self.e_ida.insert(0, "DD/MM/AAAA")
        self.e_ida.pack(side=tk.LEFT)
        self.e_ida.bind("<FocusIn>",  lambda e: self._clrph(self.e_ida,  "DD/MM/AAAA"))
        self.e_ida.bind("<FocusOut>", lambda e: self._rstph(self.e_ida,  "DD/MM/AAAA"))
        self.e_ida.bind("<KeyRelease>",
            lambda e: (self._mask_date(e, self.e_ida), self._calc_days()))
        self.lbl_ida_ok = ttk.Label(fr_ida, text="", width=3)
        self.lbl_ida_ok.pack(side=tk.LEFT, padx=(4, 0))
        self.e_hora_ida = ttk.Entry(fr_ida, width=8)
        self.e_hora_ida.insert(0, "HH:MM")
        self.e_hora_ida.pack(side=tk.LEFT, padx=(4, 0))
        self.e_hora_ida.bind("<FocusIn>",  lambda e: self._clrph(self.e_hora_ida,  "HH:MM"))
        self.e_hora_ida.bind("<FocusOut>", lambda e: self._rstph(self.e_hora_ida,  "HH:MM"))
        row += 1

        # Retorno — data e hora na mesma linha sem label intermediário
        lbl("Retorno:").grid(row=row, column=0, sticky="w", pady=4)
        fr_volta = ttk.Frame(f)
        fr_volta.grid(row=row, column=1, columnspan=4, sticky="w", pady=4)
        self.e_volta = ttk.Entry(fr_volta, width=13)
        self.e_volta.insert(0, "DD/MM/AAAA")
        self.e_volta.pack(side=tk.LEFT)
        self.e_volta.bind("<FocusIn>",  lambda e: self._clrph(self.e_volta,  "DD/MM/AAAA"))
        self.e_volta.bind("<FocusOut>", lambda e: self._rstph(self.e_volta,  "DD/MM/AAAA"))
        self.e_volta.bind("<KeyRelease>",
            lambda e: (self._mask_date(e, self.e_volta), self._calc_days()))
        self.lbl_volta_ok = ttk.Label(fr_volta, text="", width=3)
        self.lbl_volta_ok.pack(side=tk.LEFT, padx=(4, 0))
        self.e_hora_volta = ttk.Entry(fr_volta, width=8)
        self.e_hora_volta.insert(0, "HH:MM")
        self.e_hora_volta.pack(side=tk.LEFT, padx=(4, 0))
        self.e_hora_volta.bind("<FocusIn>",  lambda e: self._clrph(self.e_hora_volta,  "HH:MM"))
        self.e_hora_volta.bind("<FocusOut>", lambda e: self._rstph(self.e_hora_volta,  "HH:MM"))
        row += 1

        lbl("Dias:").grid(row=row, column=0, sticky="w", pady=3)
        self.lbl_dias = ttk.Label(f, text="—", font=(_FONT, 10))
        self.lbl_dias.grid(row=row, column=1, sticky="w")
        row += 1

        self._var_lbl_diaria = tk.StringVar(
            value=f"Valor ({self.cfg['valor_diaria']:.2f}/dia):")
        ttk.Label(f, textvariable=self._var_lbl_diaria,
                  font=(_FONT, 9, "bold")).grid(row=row, column=0, sticky="w", pady=3)
        self.lbl_val_diarias = ttk.Label(f, text="R$ 0,00",
                                         font=(_FONT, 10, "bold"), foreground=C["success"])
        self.lbl_val_diarias.grid(row=row, column=1, sticky="w")
        row += 1
        return row

    # ── Secao B: Hospedagem ───────────────────────────────────────────────────
    def _sec_hospedagem(self, f, row):
        self.var_hotel = tk.IntVar(value=0)
        lf = ttk.LabelFrame(f, text="Compare ate 3 opcoes", padding=8)
        lf.grid(row=row, column=0, columnspan=5, sticky="ew", pady=4)

        for ci, txt in enumerate(["","Nome","Link","Valor/Noite (R$)",""]):
            ttk.Label(lf, text=txt, font=("Arial", 8, "bold")).grid(
                row=0, column=ci, padx=4, pady=2)

        self.hotel_entries = []
        for i in range(3):
            r = i + 1
            ttk.Radiobutton(lf, variable=self.var_hotel, value=r).grid(
                row=r, column=0, padx=3, pady=4)
            en = ttk.Entry(lf, width=20); en.insert(0, f"Opcao {r}")
            en.grid(row=r, column=1, padx=4, pady=4)
            el = ttk.Entry(lf, width=32); el.insert(0, "https://")
            el.grid(row=r, column=2, padx=4, pady=4)
            el.bind("<FocusOut>", lambda e: self._refresh_links())
            ev = ttk.Entry(lf, width=10); ev.insert(0, "0,00")
            ev.grid(row=r, column=3, padx=4, pady=4)
            ev.bind("<FocusOut>", lambda e: self._calc_resumo())
            ModernButton(lf, text="Abrir", icon=ICONS["open"], neutral=True,
                         command=lambda idx=i: self._open_hotel_link(idx)).grid(
                row=r, column=4, padx=3)
            self.hotel_entries.append({"nome": en, "link": el, "valor": ev})

        ttk.Radiobutton(lf, text="Sem hospedagem",
                        variable=self.var_hotel, value=0).grid(
            row=4, column=0, columnspan=5, sticky="w", pady=(6, 0))
        self.var_hotel.trace_add("write", lambda *a: self._calc_resumo())
        row += 1
        return row

    # ── Secao C: Passagens ────────────────────────────────────────────────────
    def _sec_passagens(self, f, row):
        lf = ttk.LabelFrame(f, text="Bilhetes", padding=8)
        lf.grid(row=row, column=0, columnspan=5, sticky="ew", pady=4)
        lf.columnconfigure(0, weight=1)

        cols = ("Tipo","Trecho","Data/Hora","Valor (R$)","Localizador","Link")
        self.tree_pass = ttk.Treeview(lf, columns=cols, show="headings", height=4)
        ws = {"Tipo":70,"Trecho":130,"Data/Hora":105,"Valor (R$)":80,"Localizador":90,"Link":170}
        for c in cols:
            self.tree_pass.heading(c, text=c)
            self.tree_pass.column(c, width=ws[c],
                anchor="center" if c not in ("Trecho","Link") else "w")
        sb = ttk.Scrollbar(lf, orient="vertical", command=self.tree_pass.yview)
        self.tree_pass.configure(yscrollcommand=sb.set)
        self.tree_pass.grid(row=0, column=0, sticky="ew")
        sb.grid(row=0, column=1, sticky="ns")
        self.tree_pass.bind("<Double-1>", lambda e: self._edit_passagem_sel())

        bar = ttk.Frame(lf); bar.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(6,0))
        ModernButton(bar, text="Adicionar", icon=ICONS["add"],  command=lambda: self._dlg_passagem()).pack(side=tk.LEFT, padx=(0,4))
        ModernButton(bar, text="Editar",    icon=ICONS["edit"], command=self._edit_passagem_sel, neutral=True).pack(side=tk.LEFT, padx=(0,4))
        ModernButton(bar, text="Remover",   icon=ICONS["remove"], command=self._del_passagem, danger=True).pack(side=tk.LEFT, padx=(0,4))
        ModernButton(bar, text="Abrir link",icon=ICONS["open"], command=self._open_passagem_link, neutral=True).pack(side=tk.LEFT, padx=(0,4))
        self.lbl_tot_pass = ttk.Label(bar, text="Total: R$ 0,00",
                                      font=(_FONT, 9, "bold"), foreground=C["success"])
        self.lbl_tot_pass.pack(side=tk.RIGHT, padx=8)
        row += 1
        return row

    # ── Secao D: Transportes ──────────────────────────────────────────────────
    def _sec_transportes(self, f, row):
        lf = ttk.LabelFrame(f, text="Corridas e deslocamentos", padding=8)
        lf.grid(row=row, column=0, columnspan=5, sticky="ew", pady=4)
        lf.columnconfigure(0, weight=1)

        cols = ("Data","Tipo","Descricao","Valor (R$)","Anexo")
        self.tree_transp = ttk.Treeview(lf, columns=cols, show="headings", height=4)
        ws = {"Data":85,"Tipo":80,"Descricao":290,"Valor (R$)":85,"Anexo":80}
        for c in cols:
            self.tree_transp.heading(c, text=c)
            self.tree_transp.column(c, width=ws[c],
                anchor="center" if c not in ("Descricao",) else "w")
        sb = ttk.Scrollbar(lf, orient="vertical", command=self.tree_transp.yview)
        self.tree_transp.configure(yscrollcommand=sb.set)
        self.tree_transp.grid(row=0, column=0, sticky="ew")
        sb.grid(row=0, column=1, sticky="ns")
        self.tree_transp.bind("<Double-1>", lambda e: self._edit_transp_sel())

        bar = ttk.Frame(lf); bar.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(6,0))
        ModernButton(bar, text="Adicionar",  icon=ICONS["add"],    command=lambda: self._dlg_transporte()).pack(side=tk.LEFT, padx=(0,4))
        ModernButton(bar, text="Editar",     icon=ICONS["edit"],   command=self._edit_transp_sel, neutral=True).pack(side=tk.LEFT, padx=(0,4))
        ModernButton(bar, text="Remover",    icon=ICONS["remove"], command=self._del_transp, danger=True).pack(side=tk.LEFT, padx=(0,4))
        ModernButton(bar, text="Abrir anexo",icon=ICONS["attach"], command=self._open_transp_anexo, neutral=True).pack(side=tk.LEFT, padx=(0,4))
        self.lbl_tot_transp = ttk.Label(bar, text="Total: R$ 0,00",
                                        font=(_FONT, 9, "bold"), foreground=C["success"])
        self.lbl_tot_transp.pack(side=tk.RIGHT, padx=8)
        row += 1
        return row

    # ── Secao E: Materiais ────────────────────────────────────────────────────
    def _sec_materiais(self, f, row):
        lf = ttk.LabelFrame(f, text="Materiais, suprimentos e outras despesas", padding=8)
        lf.grid(row=row, column=0, columnspan=5, sticky="ew", pady=4)
        lf.columnconfigure(0, weight=1)

        cols = ("Descricao","Qtd","Valor Unit. (R$)","Total (R$)","NF Anexada")
        self.tree_mat = ttk.Treeview(lf, columns=cols, show="headings", height=4)
        ws = {"Descricao":240,"Qtd":50,"Valor Unit. (R$)":110,"Total (R$)":90,"NF Anexada":100}
        for c in cols:
            self.tree_mat.heading(c, text=c)
            self.tree_mat.column(c, width=ws[c],
                anchor="center" if c != "Descricao" else "w")
        sb = ttk.Scrollbar(lf, orient="vertical", command=self.tree_mat.yview)
        self.tree_mat.configure(yscrollcommand=sb.set)
        self.tree_mat.grid(row=0, column=0, sticky="ew")
        sb.grid(row=0, column=1, sticky="ns")
        self.tree_mat.bind("<Double-1>", lambda e: self._edit_mat_sel())

        bar = ttk.Frame(lf); bar.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(6,0))
        ModernButton(bar, text="Adicionar", icon=ICONS["add"],    command=lambda: self._dlg_material()).pack(side=tk.LEFT, padx=(0,4))
        ModernButton(bar, text="Editar",    icon=ICONS["edit"],   command=self._edit_mat_sel, neutral=True).pack(side=tk.LEFT, padx=(0,4))
        ModernButton(bar, text="Remover",   icon=ICONS["remove"], command=self._del_mat, danger=True).pack(side=tk.LEFT, padx=(0,4))
        ModernButton(bar, text="Abrir NF",  icon=ICONS["nf"],     command=self._open_mat_nf, neutral=True).pack(side=tk.LEFT, padx=(0,4))
        self.lbl_tot_mat = ttk.Label(bar, text="Total: R$ 0,00",
                                     font=(_FONT, 9, "bold"), foreground=C["success"])
        self.lbl_tot_mat.pack(side=tk.RIGHT, padx=8)
        row += 1
        return row

    # ── Secao F: Resumo ───────────────────────────────────────────────────────
    def _sec_resumo(self, f, row):
        lf = ttk.LabelFrame(f, text="Estimativa total da viagem", padding=10)
        lf.grid(row=row, column=0, columnspan=5, sticky="ew", pady=4)

        itens = [
            ("Diarias:",           "lbl_res_diarias"),
            ("Passagens:",         "lbl_res_passagens"),
            ("Transporte local:",  "lbl_res_transportes"),
            ("Hospedagem:",        "lbl_res_hotel"),
            ("Materiais:",         "lbl_res_materiais"),
        ]
        for i, (txt, attr) in enumerate(itens):
            ttk.Label(lf, text=txt).grid(row=i, column=0, sticky="w", pady=2)
            w = ttk.Label(lf, text="R$ 0,00", font=("Arial", 9))
            w.grid(row=i, column=1, sticky="w", padx=(14, 0))
            setattr(self, attr, w)

        tk.Frame(lf, bg=C["border"], height=1).grid(
            row=len(itens), column=0, columnspan=2, sticky="ew", pady=4)
        ttk.Label(lf, text="TOTAL ESTIMADO:", font=(_FONT, 10, "bold")).grid(
            row=len(itens)+1, column=0, sticky="w")
        self.lbl_res_total = ttk.Label(lf, text="R$ 0,00",
                                       font=(_FONT, 12, "bold"), foreground=C["success"])
        self.lbl_res_total.grid(row=len(itens)+1, column=1, sticky="w", padx=(14, 0))
        row += 1
        return row

    # ── Secao G: Links ────────────────────────────────────────────────────────
    def _sec_links(self, f, row):
        self._lf_links = ttk.LabelFrame(f, text="Visualizacao rapida de links", padding=8)
        self._lf_links.grid(row=row, column=0, columnspan=5, sticky="ew", pady=4)
        self._links_ctn = ttk.Frame(self._lf_links)
        self._links_ctn.pack(fill=tk.X)
        ttk.Label(self._links_ctn,
                  text="(preencha hospedagem ou passagens para ver os links aqui)",
                  font=("Arial", 8), foreground="gray").pack(anchor="w")
        row += 1
        return row

    # ── Secao H: PIX e obs ────────────────────────────────────────────────────
    def _sec_pix_obs(self, f, row):
        ttk.Label(f, text="PIX:", font=("Arial", 9, "bold")).grid(
            row=row, column=0, sticky="w", pady=4)
        self.e_pix = ttk.Entry(f, width=38)
        self.e_pix.grid(row=row, column=1, columnspan=3, sticky="w", pady=4)
        self.lbl_pix_ok = ttk.Label(f, text="", font=("Arial", 8))
        self.lbl_pix_ok.grid(row=row, column=4, sticky="w")
        self.e_pix.bind("<KeyRelease>", lambda e: self._validate_pix())
        row += 1

        ttk.Label(f, text="Observacoes:", font=("Arial", 9, "bold")).grid(
            row=row, column=0, sticky="nw", pady=4)
        self.txt_obs = tk.Text(f, height=3, width=46, font=("Arial", 9))
        self.txt_obs.grid(row=row, column=1, columnspan=4, sticky="w", pady=4)
        row += 1
        return row

    def _sec_botoes(self, f, row):
        tk.Frame(f, bg=C["border"], height=1).grid(
            row=row, column=0, columnspan=5, sticky="ew", pady=8)
        row += 1
        btns = ttk.Frame(f)
        btns.grid(row=row, column=0, columnspan=5, pady=10, sticky="w")
        ModernButton(btns, text="Gerar Canhoto", icon=ICONS["doc"], ok=True,
                     command=self._gerar_canhoto).pack(side=tk.LEFT, padx=(0, 8))
        ModernButton(btns, text="Limpar",  icon=ICONS["clear"], danger=True,
                     command=self._limpar).pack(side=tk.LEFT, padx=(0, 8))
        ModernButton(btns, text="Exportar Excel", icon=ICONS["excel"], neutral=True,
                     command=self._exportar_excel).pack(side=tk.LEFT)

    # ── Tab Painel ────────────────────────────────────────────────────────────
    def _build_painel(self):
        f = self.tab_painel

        # ── Linha de filtro ──────────────────────────────────────────────────
        flt = ttk.Frame(f)
        flt.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(flt, text="Período:", font=(_FONT, 9, "bold")).pack(side=tk.LEFT, padx=(0, 6))
        self.cmb_painel_ano = ttk.Combobox(flt, width=8, state="readonly", values=["Todos"])
        self.cmb_painel_ano.current(0)
        self.cmb_painel_ano.pack(side=tk.LEFT)
        self.cmb_painel_ano.bind("<<ComboboxSelected>>", lambda e: self._refresh_painel())
        ModernButton(flt, text="Atualizar", icon=ICONS["refresh"],
                     command=self._refresh_painel).pack(side=tk.LEFT, padx=8)

        # ── Cards ────────────────────────────────────────────────────────────
        cards_row = ttk.Frame(f)
        cards_row.pack(fill=tk.X, pady=(0, 14))
        cards_row.columnconfigure((0, 1, 2, 3, 4), weight=1)

        self._cards = {}
        card_defs = [
            ("diarias",    "Diárias",         "#4A7A9B", ICONS["doc"]),
            ("passagens",  "Passagens",        "#2E7D6B", ICONS["flight"]),
            ("transporte", "Transporte",       "#7A5C2E", ICONS["car"]),
            ("hospedagem", "Hospedagem",       "#6B2D7A", ICONS["hotel"]),
            ("materiais",  "Materiais",        "#7A2E2E", ICONS["nf"]),
            ("total",      "Total Gasto",      "#1A3A5A", ICONS["excel"]),
        ]
        for col, (key, label, color, icon) in enumerate(card_defs):
            card = tk.Frame(cards_row, bg=color, padx=14, pady=12,
                            relief="flat", bd=0)
            card.grid(row=0, column=col, padx=5, sticky="ew")
            tk.Label(card, text=f"{icon}  {label}",
                     font=(_FONT, 9), bg=color, fg="#FFFFFF").pack(anchor="w")
            val_lbl = tk.Label(card, text="R$ 0,00",
                               font=(_FONT, 15, "bold"), bg=color, fg="#FFFFFF")
            val_lbl.pack(anchor="w", pady=(4, 0))
            self._cards[key] = val_lbl

        # ── Gráfico ──────────────────────────────────────────────────────────
        chart_frame = ttk.LabelFrame(f, text="Gastos por Categoria", padding=10)
        chart_frame.pack(fill=tk.BOTH, expand=True)
        self._chart_canvas = tk.Canvas(chart_frame, bg=C["surface"],
                                       highlightthickness=0)
        self._chart_canvas.pack(fill=tk.BOTH, expand=True)
        self._chart_canvas.bind("<Configure>", lambda e: self._draw_chart())

        self._refresh_painel()

    def _refresh_painel(self):
        dados = self._load_data()

        # atualiza anos disponíveis
        anos = ["Todos"] + sorted(
            {d.get("ida", "")[-4:] for d in dados if len(d.get("ida", "")) == 10},
            reverse=True)
        self.cmb_painel_ano["values"] = anos
        if self.cmb_painel_ano.get() not in anos:
            self.cmb_painel_ano.current(0)

        # filtra por ano
        ano_sel = self.cmb_painel_ano.get()
        if ano_sel != "Todos":
            dados = [d for d in dados if d.get("ida", "")[-4:] == ano_sel]

        # totais
        t_diarias    = sum(d.get("valor",          0) for d in dados)
        t_passagens  = sum(d.get("val_passagens",  0) for d in dados)
        t_transporte = sum(d.get("val_transporte", 0) for d in dados)
        t_hospedagem = sum(d.get("val_hotel",      0) for d in dados)
        t_materiais  = sum(d.get("val_materiais",  0) for d in dados)
        t_total      = sum(d.get("val_total", d.get("valor", 0)) for d in dados)

        totais = {
            "diarias":    t_diarias,
            "passagens":  t_passagens,
            "transporte": t_transporte,
            "hospedagem": t_hospedagem,
            "materiais":  t_materiais,
            "total":      t_total,
        }
        for key, lbl in self._cards.items():
            lbl.config(text=f"R$ {totais[key]:,.2f}".replace(",", "X").replace(".", ",").replace("X", "."))

        # guarda para o gráfico
        self._chart_data = [
            ("Diárias",    t_diarias,    "#4A7A9B"),
            ("Passagens",  t_passagens,  "#2E8B6E"),
            ("Transporte", t_transporte, "#C07A2E"),
            ("Hospedagem", t_hospedagem, "#8B3DAF"),
            ("Materiais",  t_materiais,  "#AF3D3D"),
        ]
        self._draw_chart()

    def _draw_chart(self):
        cv = self._chart_canvas
        cv.delete("all")
        if not hasattr(self, "_chart_data") or not self._chart_data:
            return

        W = cv.winfo_width()
        H = cv.winfo_height()
        if W < 10 or H < 10:
            return

        pad_left = 80
        pad_right = 20
        pad_top = 20
        pad_bottom = 50
        bar_gap = 16

        labels  = [d[0] for d in self._chart_data]
        values  = [d[1] for d in self._chart_data]
        colors  = [d[2] for d in self._chart_data]
        n       = len(labels)
        max_val = max(values) if max(values) > 0 else 1

        chart_w = W - pad_left - pad_right
        chart_h = H - pad_top - pad_bottom
        bar_w   = max(10, (chart_w - bar_gap * (n + 1)) // n)

        # fundo e eixo Y
        cv.create_rectangle(pad_left, pad_top, W - pad_right, H - pad_bottom,
                             fill=C["surface"], outline=C["border"])

        # linhas guia horizontais
        for i in range(1, 6):
            y = H - pad_bottom - int(chart_h * i / 5)
            cv.create_line(pad_left, y, W - pad_right, y,
                           fill=C["border"], dash=(4, 4))
            val_ref = max_val * i / 5
            label_r = f"R${val_ref:,.0f}".replace(",", ".")
            cv.create_text(pad_left - 6, y, text=label_r,
                           anchor="e", font=(_FONT, 8), fill=C["text_dim"])

        # barras
        for i, (lbl, val, cor) in enumerate(self._chart_data):
            x0 = pad_left + bar_gap + i * (bar_w + bar_gap)
            x1 = x0 + bar_w
            bar_h_px = int(chart_h * val / max_val) if max_val > 0 else 0
            y0 = H - pad_bottom - bar_h_px
            y1 = H - pad_bottom

            # sombra (stipple para simular transparência)
            cv.create_rectangle(x0 + 3, y0 + 3, x1 + 3, y1,
                                 fill="#000000", outline="", stipple="gray25")
            # barra
            cv.create_rectangle(x0, y0, x1, y1, fill=cor, outline="", width=0)
            # destaque topo
            cv.create_rectangle(x0, y0, x1, y0 + 4, fill=_lighten(cor), outline="")

            # valor em cima
            val_str = f"R${val:,.0f}".replace(",", ".")
            cv.create_text((x0 + x1) // 2, y0 - 6,
                           text=val_str, font=(_FONT, 8, "bold"), fill=C["text"])
            # label embaixo
            cv.create_text((x0 + x1) // 2, H - pad_bottom + 14,
                           text=lbl, font=(_FONT, 9, "bold"), fill=C["text"])

    # ── Tab Historico ─────────────────────────────────────────────────────────
    def _build_historico(self):
        f = self.tab_hist

        # Filtros
        flt = ttk.LabelFrame(f, text="Filtros", padding=6)
        flt.pack(fill=tk.X, pady=(0, 8))

        ttk.Label(flt, text="Nome:", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=(0,4))
        self.e_busca_nome = ttk.Entry(flt, width=20)
        self.e_busca_nome.pack(side=tk.LEFT)
        self.e_busca_nome.bind("<KeyRelease>", lambda e: self._filter_hist())

        ttk.Label(flt, text="Destino:", font=("Arial", 9, "bold")).pack(side=tk.LEFT, padx=(8,4))
        self.e_busca_dest = ttk.Entry(flt, width=16)
        self.e_busca_dest.pack(side=tk.LEFT)
        self.e_busca_dest.bind("<KeyRelease>", lambda e: self._filter_hist())

        ttk.Label(flt, text="Mes:").pack(side=tk.LEFT, padx=(8, 4))
        self.cmb_mes = ttk.Combobox(flt, width=10, state="readonly", values=MESES)
        self.cmb_mes.current(0)
        self.cmb_mes.pack(side=tk.LEFT)
        self.cmb_mes.bind("<<ComboboxSelected>>", lambda e: self._filter_hist())

        ttk.Label(flt, text="Ano:").pack(side=tk.LEFT, padx=(6, 4))
        self.cmb_ano = ttk.Combobox(flt, width=6, state="readonly", values=["Todos"])
        self.cmb_ano.current(0)
        self.cmb_ano.pack(side=tk.LEFT)
        self.cmb_ano.bind("<<ComboboxSelected>>", lambda e: self._filter_hist())

        ModernButton(flt, text="Atualizar", icon=ICONS["refresh"],
                     command=self._load_hist, neutral=True).pack(side=tk.LEFT, padx=8)
        ModernButton(flt, text="Excel",  icon=ICONS["excel"],
                     command=self._exportar_excel, neutral=True).pack(side=tk.RIGHT)
        ModernButton(flt, text="Excluir", icon=ICONS["remove"],
                     command=self._del_registro, danger=True).pack(side=tk.RIGHT, padx=4)
        ModernButton(flt, text="Editar",  icon=ICONS["edit"],
                     command=lambda: self._open_edit_dlg()).pack(side=tk.RIGHT, padx=4)

        # Treeview
        cols = ("Nome","Destino","Partida","Retorno","Hotel","Pass.","Mat.","Dias","Total (R$)")
        self.tree_hist = ttk.Treeview(f, columns=cols, show="headings", height=16)
        ws = {"Nome":130,"Destino":110,"Partida":98,"Retorno":98,
              "Hotel":82,"Pass.":45,"Mat.":45,"Dias":42,"Total (R$)":88}
        ctr = {"Partida","Retorno","Hotel","Pass.","Mat.","Dias","Total (R$)"}
        for c in cols:
            self.tree_hist.heading(c, text=c)
            self.tree_hist.column(c, width=ws[c],
                anchor="center" if c in ctr else "w")
        vsb = ttk.Scrollbar(f, orient="vertical", command=self.tree_hist.yview)
        self.tree_hist.configure(yscrollcommand=vsb.set)
        self.tree_hist.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        self.tree_hist.bind("<Double-1>", lambda e: self._open_edit_dlg())

        self.lbl_hist_total = ttk.Label(f, text="", font=(_FONT, 9, "bold"))
        self.lbl_hist_total.pack(anchor="e", pady=4)
        self._load_hist()

    # ── Settings lock ─────────────────────────────────────────────────────────
    def _build_cfg_lock(self):
        """Mostra tela de login ou cadastro de senha na aba Configuracoes."""
        for w in self.tab_cfg.winfo_children():
            w.destroy()
        self._cfg_unlocked = False

        has_creds = bool(self.cfg.get("cfg_hash"))
        title_txt = "Acesso as Configuracoes" if has_creds else "Criar Acesso"

        frm = ttk.Frame(self.tab_cfg, padding=30)
        frm.place(relx=0.5, rely=0.35, anchor="center")

        tk.Label(frm, text=title_txt,
                 font=("Arial", 14, "bold"), fg="#2563EB").grid(
            row=0, column=0, columnspan=2, pady=(0, 18))

        ttk.Label(frm, text="Usuario:", font=("Arial", 10)).grid(
            row=1, column=0, sticky="w", pady=6)
        e_user = ttk.Entry(frm, width=26)
        e_user.grid(row=1, column=1, sticky="w", pady=6, padx=(8, 0))

        ttk.Label(frm, text="Senha:", font=("Arial", 10)).grid(
            row=2, column=0, sticky="w", pady=6)
        e_pass = ttk.Entry(frm, width=26, show="*")
        e_pass.grid(row=2, column=1, sticky="w", pady=6, padx=(8, 0))
        e_pass.bind("<Return>", lambda e: btn_cmd())

        msg_var = tk.StringVar()
        ttk.Label(frm, textvariable=msg_var,
                  foreground="red", font=("Arial", 9)).grid(
            row=3, column=0, columnspan=2, pady=4)

        def btn_cmd():
            user = e_user.get().strip()
            pwd  = e_pass.get().strip()
            if not user or not pwd:
                msg_var.set("Preencha usuario e senha.")
                return
            if not has_creds:
                h, salt = _hash_senha(user + pwd)
                self.cfg["cfg_user"] = user
                self.cfg["cfg_hash"] = h
                self.cfg["cfg_salt"] = salt
                self._save_cfg()
                self._cfg_senha = pwd
                self._unlock_cfg()
            else:
                salt = self.cfg.get("cfg_salt", "")
                # compatibilidade com hashes antigos (SHA-256 sem salt)
                if not salt:
                    import hashlib as _hl
                    ok = _hl.sha256((user + pwd).strip().encode()).hexdigest() == self.cfg["cfg_hash"]
                    if ok:
                        # migra para scrypt na primeira autenticação bem-sucedida
                        h, salt = _hash_senha(user + pwd)
                        self.cfg["cfg_hash"] = h
                        self.cfg["cfg_salt"] = salt
                        self._save_cfg()
                else:
                    ok = _verificar_senha(user + pwd, self.cfg["cfg_hash"], salt)
                if ok:
                    self._cfg_senha = pwd
                    self._unlock_cfg()
                else:
                    msg_var.set("Usuario ou senha incorretos.")
                    e_pass.delete(0, tk.END)

        btn_lbl = "Entrar" if has_creds else "Criar acesso"
        btn_icon = ICONS["key"] if has_creds else ICONS["save"]
        ModernButton(frm, text=btn_lbl, icon=btn_icon, ok=True,
                     command=btn_cmd).grid(row=4, column=0, columnspan=2, pady=12)

        if not has_creds:
            ttk.Label(frm,
                text="Defina as credenciais de acesso as configuracoes.",
                font=("Arial", 8), foreground="gray").grid(
                row=5, column=0, columnspan=2)

    def _unlock_cfg(self):
        self._cfg_unlocked = True
        for w in self.tab_cfg.winfo_children():
            w.destroy()
        self._build_configuracoes()

    # ── Tab Configuracoes (desbloqueada) ──────────────────────────────────────
    def _build_configuracoes(self):
        f = self.tab_cfg

        # Header da aba
        hdr = ttk.Frame(f)
        hdr.pack(fill=tk.X, pady=(0, 10))
        tk.Label(hdr, text=f"  Configuracoes  (usuario: {self.cfg.get('cfg_user','')})",
                 font=("Arial", 10, "bold"), fg="#2563EB").pack(side=tk.LEFT)
        ModernButton(hdr, text="Bloquear",     icon=ICONS["lock"],  danger=True,
                     command=self._build_cfg_lock).pack(side=tk.RIGHT)
        ModernButton(hdr, text="Alterar senha",icon=ICONS["key"],   neutral=True,
                     command=self._dlg_alterar_senha).pack(side=tk.RIGHT, padx=6)

        nb_cfg = ttk.Notebook(f)
        nb_cfg.pack(fill=tk.BOTH, expand=True)

        t1 = ttk.Frame(nb_cfg, padding=12)
        t2 = ttk.Frame(nb_cfg, padding=12)
        t3 = ttk.Frame(nb_cfg, padding=12)
        nb_cfg.add(t1, text="  Geral  ")
        nb_cfg.add(t2, text="  E-mail  ")
        nb_cfg.add(t3, text="  Sistema  ")

        # ── Geral
        lf_d = ttk.LabelFrame(t1, text="Valor da Diaria", padding=10)
        lf_d.pack(fill=tk.X, pady=(0, 10))
        ttk.Label(lf_d, text="Valor por dia (R$):").pack(side=tk.LEFT, padx=(0, 8))
        self.e_diaria = ttk.Entry(lf_d, width=12)
        self.e_diaria.insert(0, f"{self.cfg['valor_diaria']:.2f}")
        self.e_diaria.pack(side=tk.LEFT)
        ModernButton(lf_d, text="Aplicar", icon=ICONS["save"], ok=True,
                     command=self._apply_diaria).pack(side=tk.LEFT, padx=8)

        lf_b = ttk.LabelFrame(t1, text="Backup", padding=10)
        lf_b.pack(fill=tk.X, pady=(0, 10))
        self.var_backup = tk.BooleanVar(value=self.cfg.get("backup_automatico", True))
        ttk.Checkbutton(lf_b,
            text="Criar backup diario automatico ao iniciar",
            variable=self.var_backup,
            command=lambda: (self.cfg.update(backup_automatico=self.var_backup.get()),
                             self._save_cfg())).pack(anchor="w")
        ModernButton(lf_b, text="Fazer backup agora", icon=ICONS["backup"], neutral=True,
                     command=self._backup_manual).pack(anchor="w", pady=(8, 0))

        lf_t = ttk.LabelFrame(t1, text="Aparencia", padding=10)
        lf_t.pack(fill=tk.X)
        ttk.Label(lf_t, text="Tema:").pack(side=tk.LEFT, padx=(0, 8))
        self.var_tema_cfg = tk.StringVar(value=self.cfg["tema"])
        ttk.Radiobutton(lf_t, text=f"{ICONS['light']}  Claro",  variable=self.var_tema_cfg,
                        value="light", command=self._apply_tema_cfg).pack(side=tk.LEFT, padx=4)
        ttk.Radiobutton(lf_t, text=f"{ICONS['dark']}  Escuro",  variable=self.var_tema_cfg,
                        value="dark",  command=self._apply_tema_cfg).pack(side=tk.LEFT, padx=4)

        # ── E-mail
        lf_e = ttk.LabelFrame(t2, text="SMTP", padding=10)
        lf_e.pack(fill=tk.X, pady=(0, 10))

        self.var_auto_email = tk.BooleanVar(value=self.cfg.get("auto_email", False))
        ttk.Checkbutton(lf_e,
            text="Enviar e-mail automaticamente ao gerar canhoto",
            variable=self.var_auto_email).grid(
            row=0, column=0, columnspan=2, sticky="w", pady=(0, 8))

        campos = [
            ("Remetente:",    "email_remetente",    False, 34),
            ("Senha/App Pass:","email_senha",         True,  34),
            ("SMTP:",          "email_smtp",          False, 22),
            ("Porta:",         "email_porta",         False, 7),
            ("Destinatario:",  "email_destinatario",  False, 34),
        ]
        self._email_entries = {}
        for i, (lbl, chave, sec, w) in enumerate(campos):
            ttk.Label(lf_e, text=lbl).grid(row=i+1, column=0, sticky="w", pady=3, padx=(0,8))
            e = ttk.Entry(lf_e, width=w, show="*" if sec else "")
            e.insert(0, str(self.cfg.get(chave, "")))
            e.grid(row=i+1, column=1, sticky="w", pady=3)
            self._email_entries[chave] = e

        n = len(campos)
        ttk.Label(lf_e,
            text="Gmail/Workspace: use Senha de App (myaccount.google.com > Senhas de app)",
            font=("Arial", 8), foreground="gray").grid(
            row=n+1, column=0, columnspan=2, sticky="w", pady=(4, 6))

        btns_e = ttk.Frame(lf_e)
        btns_e.grid(row=n+2, column=0, columnspan=2, sticky="w")
        ModernButton(btns_e, text="Salvar configuracoes", icon=ICONS["save"], ok=True,
                     command=self._save_email_cfg).pack(side=tk.LEFT, padx=(0, 8))
        ModernButton(btns_e, text="Enviar e-mail de teste", icon=ICONS["email"], neutral=True,
                     command=lambda: self._enviar_auto(
                       "[TESTE] Configuracao OK — Registro de Viagens",
                       "[TESTE] Se recebeu este e-mail, a configuracao esta correta."
                   )).pack(side=tk.LEFT)

        # ── Sistema
        lf_s = ttk.LabelFrame(t3, text="Informacoes do sistema", padding=10)
        lf_s.pack(fill=tk.X)
        for linha in [
            f"Arquivo de dados : {os.path.abspath(DATA_FILE)}",
            f"Backups          : {os.path.abspath(BACKUP_DIR)}",
            f"Anexos           : {os.path.abspath(ATTACH_DIR)}",
            f"sv_ttk           : {'Instalado' if _SV_TTK else 'Nao instalado (pip install sv-ttk)'}",
            f"ReportLab        : {'Instalado' if _REPORTLAB else 'Nao instalado (pip install reportlab)'}",
        ]:
            ttk.Label(lf_s, text=linha, font=("Courier", 9)).pack(anchor="w")

    def _dlg_alterar_senha(self):
        win = tk.Toplevel(self); win.title("Alterar Senha"); win.geometry("360x220"); win.grab_set()
        f = ttk.Frame(win, padding=20); f.pack(fill=tk.BOTH, expand=True)
        ttk.Label(f, text="Alterar credenciais de acesso",
                  font=("Arial", 11, "bold")).grid(row=0, column=0, columnspan=2, pady=(0,12))
        labels = ["Senha atual:", "Novo usuario:", "Nova senha:", "Confirmar senha:"]
        entries = {}
        for i, lbl in enumerate(labels):
            ttk.Label(f, text=lbl).grid(row=i+1, column=0, sticky="w", pady=4)
            e = ttk.Entry(f, width=24, show="*")
            e.grid(row=i+1, column=1, sticky="w", pady=4, padx=(8,0))
            entries[lbl] = e
        msg = tk.StringVar()
        ttk.Label(f, textvariable=msg, foreground="red", font=("Arial",8)).grid(
            row=5, column=0, columnspan=2)
        def _salvar():
            user_atual  = self.cfg.get("cfg_user", "")
            senha_atual = entries["Senha atual:"].get().strip()
            salt_atual  = self.cfg.get("cfg_salt", "")
            if salt_atual:
                ok = _verificar_senha(user_atual + senha_atual, self.cfg["cfg_hash"], salt_atual)
            else:
                import hashlib as _hl
                ok = _hl.sha256((user_atual + senha_atual).strip().encode()).hexdigest() == self.cfg["cfg_hash"]
            if not ok:
                msg.set("Senha atual incorreta.")
                return
            novo_user  = entries["Novo usuario:"].get().strip()
            nova_senha = entries["Nova senha:"].get().strip()
            confirma   = entries["Confirmar senha:"].get().strip()
            if nova_senha != confirma:
                msg.set("As senhas nao coincidem.")
                return
            if not novo_user or not nova_senha:
                msg.set("Preencha todos os campos.")
                return
            h, salt = _hash_senha(novo_user + nova_senha)
            self.cfg["cfg_user"] = novo_user
            self.cfg["cfg_hash"] = h
            self.cfg["cfg_salt"] = salt
            # re-cifra senha de email com nova chave
            if _CRYPTO and self.cfg.get("email_senha"):
                email_plain = _decifrar(self.cfg["email_senha"], senha_atual, salt_atual or salt)
                self.cfg["email_senha"] = _cifrar(email_plain, nova_senha, salt)
            self._cfg_senha = nova_senha
            self._save_cfg()
            messagebox.showinfo("Salvo", "Credenciais atualizadas!", parent=win)
            win.destroy()
        ModernButton(f, text="Salvar", icon=ICONS["save"], ok=True,
                     command=_salvar).grid(row=6, column=0, columnspan=2, pady=10)

    def _apply_diaria(self):
        try:
            v = float(self.e_diaria.get().strip().replace(",","."))
            self.cfg["valor_diaria"] = v
            self._save_cfg()
            self._var_lbl_diaria.set(f"Valor ({v:.2f}/dia):")
            self._calc_days()
            messagebox.showinfo("OK", f"Diaria atualizada: R$ {v:.2f}")
        except ValueError:
            messagebox.showerror("Erro", "Valor invalido.")

    def _apply_tema_cfg(self):
        novo = self.var_tema_cfg.get()
        self.cfg["tema"] = novo
        self._save_cfg()
        self._apply_theme(novo)
        prox = TEMA_CICLO[(TEMA_CICLO.index(novo) + 1) % len(TEMA_CICLO)]
        self.btn_tema.config(text=TEMA_LABELS[prox],
                             bg=C["btn_neutral"], activebackground=C["btn_neutral_h"])

    def _save_email_cfg(self):
        for chave, entry in self._email_entries.items():
            val = entry.get().strip()
            if chave == "email_porta":
                self.cfg[chave] = int(val) if val.isdigit() else 587
            elif chave == "email_senha":
                salt = self.cfg.get("cfg_salt", "")
                self.cfg[chave] = _cifrar(val, self._cfg_senha, salt) if (_CRYPTO and salt) else val
            else:
                self.cfg[chave] = val
        self.cfg["auto_email"] = self.var_auto_email.get()
        self._save_cfg()
        aviso = "" if _CRYPTO else "\n(instale 'cryptography' para cifrar a senha)"
        status = ("Salvo. Envio automatico ATIVO." if self.cfg["auto_email"] else "Salvo.") + aviso
        messagebox.showinfo("Salvo", status)

    def _get_email_senha(self) -> str:
        """Retorna a senha de e-mail decifrada."""
        raw  = self.cfg.get("email_senha", "")
        salt = self.cfg.get("cfg_salt", "")
        if _CRYPTO and salt and self._cfg_senha:
            return _decifrar(raw, self._cfg_senha, salt)
        return raw

    # ── Eventos gerais ────────────────────────────────────────────────────────
    def _on_tab(self, _e):
        idx = self.nb.index(self.nb.select())
        if idx == 1:
            self._load_hist()
        elif idx == 2:
            self._refresh_painel()
        elif idx == 3 and not self._cfg_unlocked:
            self._build_cfg_lock()

    def _clrph(self, e, ph):
        if e.get() == ph: e.delete(0, tk.END)

    def _rstph(self, e, ph):
        if not e.get().strip(): e.insert(0, ph)

    def _mask_date(self, event, entry):
        if event.keysym in ("BackSpace","Delete"): return
        d = re.sub(r"\D","", entry.get())[:8]
        if   len(d) <= 2: v = d
        elif len(d) <= 4: v = f"{d[:2]}/{d[2:]}"
        else:             v = f"{d[:2]}/{d[2:4]}/{d[4:]}"
        if v != entry.get():
            entry.delete(0, tk.END); entry.insert(0, v)

    def _calc_days(self):
        ida   = self.e_ida.get().strip()
        volta = self.e_volta.get().strip()
        self._validate_date_lbl(ida,   self.lbl_ida_ok)
        self._validate_date_lbl(volta, self.lbl_volta_ok)
        dias = self._days(ida, volta)
        vd   = self.cfg["valor_diaria"]
        if dias > 0:
            self.lbl_dias.config(text=f"{dias} dia(s)")
            self.lbl_val_diarias.config(text=f"R$ {dias*vd:.2f}")
        else:
            self.lbl_dias.config(text="—")
            self.lbl_val_diarias.config(text="R$ 0,00")
        self._calc_resumo()

    def _validate_date_lbl(self, v, lbl):
        if len(v) == 10:
            try:
                datetime.strptime(v, "%d/%m/%Y")
                lbl.config(text="ok", foreground="#16A34A")
            except ValueError:
                lbl.config(text="x",  foreground="#DC2626")
        else:
            lbl.config(text="")

    def _validate_pix(self):
        pix = self.e_pix.get().strip()
        if not pix: self.lbl_pix_ok.config(text=""); return
        d = re.sub(r"\D","", pix)
        if   re.match(r"^[\w\.\+\-]+@[\w\.-]+\.\w+$", pix): self.lbl_pix_ok.config(text="e-mail",    foreground="#16A34A")
        elif len(d) == 11:                                    self.lbl_pix_ok.config(text="CPF",       foreground="#16A34A")
        elif len(d) == 14:                                    self.lbl_pix_ok.config(text="CNPJ",      foreground="#16A34A")
        elif len(d) in (10,11):                               self.lbl_pix_ok.config(text="Telefone",  foreground="#16A34A")
        else:                                                 self.lbl_pix_ok.config(text="Aleatoria?",foreground="#D97706")

    def _open_hotel_link(self, idx):
        link = self.hotel_entries[idx]["link"].get().strip()
        if link and link not in ("https://","http://") and link.startswith("http"):
            webbrowser.open(link)
        else:
            messagebox.showwarning("Aviso", "Link invalido.")

    # ── Calculo geral ─────────────────────────────────────────────────────────
    def _days(self, a, b) -> int:
        try:
            delta = (datetime.strptime(b, "%d/%m/%Y") -
                     datetime.strptime(a, "%d/%m/%Y")).days + 1
            return delta if delta >= 1 else 0
        except ValueError:
            return 0

    def _get_hotel(self):
        s = self.var_hotel.get()
        if s == 0:
            return {"selecionado":"Nao","nome":"","link":"","valor_noite":0.0,"todas_opcoes":[]}
        h = self.hotel_entries[s-1]
        try: val = float(h["valor"].get().strip().replace(",",".").replace("R$",""))
        except ValueError: val = 0.0
        todas = []
        for i, he in enumerate(self.hotel_entries):
            nome = he["nome"].get().strip()
            link = he["link"].get().strip()
            try: v = float(he["valor"].get().strip().replace(",",".").replace("R$",""))
            except ValueError: v = 0.0
            if nome and nome not in (f"Opcao {i+1}",):
                todas.append({"opcao": i+1, "nome": nome, "link": link, "valor_noite": v,
                               "selecionada": (i+1 == s)})
        return {"selecionado":"Sim","nome":h["nome"].get().strip(),
                "link":h["link"].get().strip(),"valor_noite":val,"todas_opcoes":todas}

    def _calc_resumo(self):
        dias  = self._days(self.e_ida.get().strip(), self.e_volta.get().strip())
        vd    = self.cfg["valor_diaria"]
        val_d = dias * vd if dias > 0 else 0
        val_p = sum(p.get("valor",0) for p in self._passagens)
        val_t = sum(t.get("valor",0) for t in self._transportes)
        val_m = sum(m.get("total",0) for m in self._materiais)
        s     = self.var_hotel.get()
        val_h = 0.0
        if s > 0 and dias > 0:
            try:
                val_h = float(
                    self.hotel_entries[s-1]["valor"].get()
                    .strip().replace(",",".").replace("R$","")) * dias
            except ValueError: val_h = 0.0

        self.lbl_res_diarias.config(    text=f"R$ {val_d:.2f}" + (f" ({dias}d)" if dias>0 else ""))
        self.lbl_res_passagens.config(  text=f"R$ {val_p:.2f}" + (f" ({len(self._passagens)} bilhete(s))" if self._passagens else ""))
        self.lbl_res_transportes.config(text=f"R$ {val_t:.2f}" + (f" ({len(self._transportes)} corrida(s))" if self._transportes else ""))
        self.lbl_res_hotel.config(      text=f"R$ {val_h:.2f}")
        self.lbl_res_materiais.config(  text=f"R$ {val_m:.2f}" + (f" ({len(self._materiais)} item(ns))" if self._materiais else ""))
        self.lbl_res_total.config(      text=f"R$ {val_d+val_p+val_t+val_h+val_m:.2f}")

    # ── Passagens CRUD ────────────────────────────────────────────────────────
    def _dlg_passagem(self, idx=None):
        ex  = self._passagens[idx] if idx is not None else None
        editando = idx is not None

        win = tk.Toplevel(self)
        win.title("Passagem" if editando else "Adicionar Passagem")
        win.geometry("520x460")
        win.grab_set()

        # ── Scroll interno
        canvas = tk.Canvas(win, highlightthickness=0)
        vsb    = ttk.Scrollbar(win, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vsb.set)
        vsb.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        f = ttk.Frame(canvas, padding=15)
        wid = canvas.create_window((0,0), window=f, anchor="nw")
        f.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(wid, width=e.width-4))

        # ── Tipo e toggle ida/volta
        ttk.Label(f, text="Tipo:", font=("Arial",9,"bold")).grid(row=0, column=0, sticky="w", pady=4)
        cmb = ttk.Combobox(f, values=TIPOS_PASSAGEM, state="readonly", width=18)
        cmb.set(ex.get("tipo","Aereo") if ex else "Aereo")
        cmb.grid(row=0, column=1, sticky="w")

        var_rt = tk.BooleanVar(value=False)
        if not editando:
            chk_rt = ttk.Checkbutton(f, text="Ida e Volta", variable=var_rt,
                                      command=lambda: _toggle_volta())
            chk_rt.grid(row=0, column=2, columnspan=2, sticky="w", padx=(14,0))

        # ── Separador IDA
        tk.Label(f, text=" IDA ", font=("Arial",9,"bold"),
                 bg="#2563EB", fg="white", padx=6, pady=2).grid(
            row=1, column=0, columnspan=4, sticky="w", pady=(8,2))

        ttk.Label(f, text="Trecho:", font=("Arial",9,"bold")).grid(row=2, column=0, sticky="w", pady=3)
        e_tr = ttk.Entry(f, width=30)
        e_tr.insert(0, ex.get("trecho","") if ex else "")
        e_tr.grid(row=2, column=1, columnspan=3, sticky="w")
        ttk.Label(f, text="ex: BSB > GRU", font=("Arial",8), foreground="gray").grid(
            row=2, column=4, sticky="w", padx=4)

        ttk.Label(f, text="Data:", font=("Arial",9,"bold")).grid(row=3, column=0, sticky="w", pady=3)
        e_dt = ttk.Entry(f, width=13); e_dt.insert(0, ex.get("data","") if ex else "")
        e_dt.grid(row=3, column=1, sticky="w")
        e_dt.bind("<KeyRelease>", lambda e: self._mask_date(e, e_dt))
        ttk.Label(f, text="Hora:").grid(row=3, column=2, sticky="e", padx=(10,5))
        e_hr = ttk.Entry(f, width=8); e_hr.insert(0, ex.get("hora","") if ex else "")
        e_hr.grid(row=3, column=3, sticky="w")

        ttk.Label(f, text="Valor (R$):", font=("Arial",9,"bold")).grid(row=4, column=0, sticky="w", pady=3)
        e_vl = ttk.Entry(f, width=14)
        e_vl.insert(0, f"{ex['valor']:.2f}" if ex else "0,00")
        e_vl.grid(row=4, column=1, sticky="w")

        ttk.Label(f, text="Localizador:").grid(row=5, column=0, sticky="w", pady=3)
        e_lc = ttk.Entry(f, width=18); e_lc.insert(0, ex.get("localizador","") if ex else "")
        e_lc.grid(row=5, column=1, sticky="w")

        ttk.Label(f, text="Link:").grid(row=6, column=0, sticky="w", pady=3)
        e_lk = ttk.Entry(f, width=38)
        e_lk.insert(0, ex.get("link","https://") if ex else "https://")
        e_lk.grid(row=6, column=1, columnspan=3, sticky="w")

        # ── Separador VOLTA (oculto inicialmente)
        lbl_sep_v = tk.Label(f, text=" VOLTA ", font=("Arial",9,"bold"),
                              bg="#7C3AED", fg="white", padx=6, pady=2)
        lbl_hint_v = ttk.Label(f, text="Trecho inverso preenchido automaticamente",
                                font=("Arial",8), foreground="gray")

        ttk.Label(f, text="Trecho:").grid(row=8, column=0, sticky="w", pady=3)
        e_tr_v = ttk.Entry(f, width=30); e_tr_v.grid(row=8, column=1, columnspan=3, sticky="w")

        ttk.Label(f, text="Data:").grid(row=9, column=0, sticky="w", pady=3)
        e_dt_v = ttk.Entry(f, width=13); e_dt_v.grid(row=9, column=1, sticky="w")
        e_dt_v.bind("<KeyRelease>", lambda e: self._mask_date(e, e_dt_v))
        ttk.Label(f, text="Hora:").grid(row=9, column=2, sticky="e", padx=(10,5))
        e_hr_v = ttk.Entry(f, width=8); e_hr_v.grid(row=9, column=3, sticky="w")

        ttk.Label(f, text="Valor (R$):").grid(row=10, column=0, sticky="w", pady=3)
        e_vl_v = ttk.Entry(f, width=14); e_vl_v.insert(0, "0,00")
        e_vl_v.grid(row=10, column=1, sticky="w")

        ttk.Label(f, text="Localizador:").grid(row=11, column=0, sticky="w", pady=3)
        e_lc_v = ttk.Entry(f, width=18); e_lc_v.grid(row=11, column=1, sticky="w")

        ttk.Label(f, text="Link:").grid(row=12, column=0, sticky="w", pady=3)
        e_lk_v = ttk.Entry(f, width=38); e_lk_v.insert(0, "https://")
        e_lk_v.grid(row=12, column=1, columnspan=3, sticky="w")

        # Widgets da secao volta — comecem ocultos
        volta_widgets = [lbl_sep_v, lbl_hint_v,
                         e_tr_v, e_dt_v, e_hr_v, e_vl_v, e_lc_v, e_lk_v]
        _rows_volta = [7, 7, 8, 9, 9, 10, 11, 12]  # row de cada widget
        for w in volta_widgets:
            w.grid_remove()

        def _toggle_volta():
            if var_rt.get():
                lbl_sep_v.grid(row=7, column=0, columnspan=4, sticky="w", pady=(10,2))
                lbl_hint_v.grid(row=7, column=4, sticky="w", padx=4)
                for w in [e_tr_v, e_dt_v, e_hr_v, e_vl_v, e_lc_v, e_lk_v]:
                    w.grid()
                win.geometry("520x640")
                # Auto-preencher trecho inverso
                trecho = e_tr.get().strip()
                if ">" in trecho:
                    partes = [p.strip() for p in trecho.split(">")]
                    if len(partes) == 2:
                        e_tr_v.delete(0, tk.END)
                        e_tr_v.insert(0, f"{partes[1]} > {partes[0]}")
                # Auto-preencher data de volta com data de retorno da viagem
                data_ret = self.e_volta.get().strip()
                if len(data_ret) == 10:
                    e_dt_v.delete(0, tk.END)
                    e_dt_v.insert(0, data_ret)
                # Mesmo link por padrao
                link_ida = e_lk.get().strip()
                if link_ida not in ("https://", "http://", ""):
                    e_lk_v.delete(0, tk.END)
                    e_lk_v.insert(0, link_ida)
            else:
                for w in volta_widgets:
                    w.grid_remove()
                win.geometry("520x460")

        # Atualizar trecho volta quando trecho ida muda
        def _on_trecho_change(*_):
            if var_rt.get() and ">" in e_tr.get():
                partes = [p.strip() for p in e_tr.get().split(">")]
                if len(partes) == 2:
                    e_tr_v.delete(0, tk.END)
                    e_tr_v.insert(0, f"{partes[1]} > {partes[0]}")
        e_tr.bind("<KeyRelease>", _on_trecho_change)

        bf = ttk.Frame(f)
        bf.grid(row=13, column=0, columnspan=5, pady=14, sticky="w")

        def salvar():
            try: v = float(e_vl.get().strip().replace(",","."))
            except ValueError: v = 0.0
            entry_ida = {
                "tipo":        cmb.get(),
                "trecho":      e_tr.get().strip(),
                "data":        e_dt.get().strip(),
                "hora":        e_hr.get().strip(),
                "valor":       v,
                "localizador": e_lc.get().strip(),
                "link":        e_lk.get().strip(),
                "sentido":     "Ida",
            }
            if editando:
                self._passagens[idx] = entry_ida
            else:
                self._passagens.append(entry_ida)
                if var_rt.get():
                    try: vv = float(e_vl_v.get().strip().replace(",","."))
                    except ValueError: vv = 0.0
                    entry_volta = {
                        "tipo":        cmb.get(),
                        "trecho":      e_tr_v.get().strip(),
                        "data":        e_dt_v.get().strip(),
                        "hora":        e_hr_v.get().strip(),
                        "valor":       vv,
                        "localizador": e_lc_v.get().strip(),
                        "link":        e_lk_v.get().strip(),
                        "sentido":     "Volta",
                    }
                    self._passagens.append(entry_volta)
            self._sync_pass()
            self._calc_resumo()
            self._refresh_links()
            win.destroy()

        ModernButton(bf, text="Salvar",   icon=ICONS["save"],   ok=True,      command=salvar).pack(side=tk.LEFT, padx=5)
        ModernButton(bf, text="Cancelar", icon=ICONS["cancel"], danger=True,  command=win.destroy).pack(side=tk.LEFT)

    def _edit_passagem_sel(self):
        s = self.tree_pass.focus()
        if s:
            try: self._dlg_passagem(idx=int(s))
            except (ValueError,IndexError): pass

    def _del_passagem(self):
        s = self.tree_pass.focus()
        if not s: return
        try: self._passagens.pop(int(s))
        except (ValueError,IndexError): return
        self._sync_pass(); self._calc_resumo(); self._refresh_links()

    def _open_passagem_link(self):
        s = self.tree_pass.focus()
        if not s: return
        try:
            lk = self._passagens[int(s)].get("link","")
            if lk and lk.startswith("http"): webbrowser.open(lk)
            else: messagebox.showwarning("Aviso","Link nao preenchido.")
        except (ValueError,IndexError): pass

    def _sync_pass(self):
        self.tree_pass.delete(*self.tree_pass.get_children())
        for i, p in enumerate(self._passagens):
            dh      = p.get("data","") + (f" {p['hora']}" if p.get("hora") else "")
            tipo_ex = p.get("tipo","")
            sentido = p.get("sentido","")
            if sentido:
                tipo_ex = f"{tipo_ex} ({sentido})"
            self.tree_pass.insert("", tk.END, iid=str(i), values=(
                tipo_ex, p.get("trecho",""), dh,
                f"R$ {p.get('valor',0):.2f}", p.get("localizador",""), p.get("link","")))
        self.lbl_tot_pass.config(
            text=f"Total: R$ {sum(p.get('valor',0) for p in self._passagens):.2f}")

    # ── Transportes CRUD ──────────────────────────────────────────────────────
    def _dlg_transporte(self, idx=None):
        ex  = self._transportes[idx] if idx is not None else None
        win = tk.Toplevel(self); win.title("Transporte"); win.geometry("430x260"); win.grab_set()
        f   = ttk.Frame(win, padding=15); f.pack(fill=tk.BOTH, expand=True)
        ttk.Label(f, text="Data:").grid(row=0, column=0, sticky="w", pady=3)
        e_dt = ttk.Entry(f, width=13); e_dt.insert(0, ex.get("data","") if ex else "")
        e_dt.grid(row=0, column=1, sticky="w")
        e_dt.bind("<KeyRelease>", lambda e: self._mask_date(e, e_dt))
        ttk.Label(f, text="Tipo:").grid(row=1, column=0, sticky="w", pady=3)
        cmb = ttk.Combobox(f, values=TIPOS_TRANSPORTE, state="readonly", width=18)
        cmb.set(ex.get("tipo","Uber") if ex else "Uber"); cmb.grid(row=1, column=1, sticky="w")
        ttk.Label(f, text="Descricao:").grid(row=2, column=0, sticky="w", pady=3)
        e_dc = ttk.Entry(f, width=32); e_dc.insert(0, ex.get("descricao","") if ex else "")
        e_dc.grid(row=2, column=1, columnspan=2, sticky="w")
        ttk.Label(f, text="Valor (R$):").grid(row=3, column=0, sticky="w", pady=3)
        e_vl = ttk.Entry(f, width=14); e_vl.insert(0, f"{ex['valor']:.2f}" if ex else "0,00")
        e_vl.grid(row=3, column=1, sticky="w")
        # Anexo
        ttk.Label(f, text="Anexo:").grid(row=4, column=0, sticky="w", pady=3)
        e_anx = ttk.Entry(f, width=28); e_anx.insert(0, ex.get("anexo","") if ex else "")
        e_anx.grid(row=4, column=1, sticky="w")
        def _browse():
            fp = filedialog.askopenfilename(
                title="Selecionar comprovante",
                filetypes=[("PDF/Imagem","*.pdf *.png *.jpg *.jpeg *.bmp"),("Todos","*.*")])
            if fp: e_anx.delete(0,tk.END); e_anx.insert(0,fp)
        ModernButton(f, text="Selecionar", icon=ICONS["attach"], neutral=True,
                     command=_browse).grid(row=4, column=2, padx=4)
        def salvar():
            try: v = float(e_vl.get().strip().replace(",","."))
            except ValueError: v = 0.0
            entry = {"data":e_dt.get().strip(),"tipo":cmb.get(),
                     "descricao":e_dc.get().strip(),"valor":v,
                     "anexo":e_anx.get().strip()}
            if idx is not None: self._transportes[idx] = entry
            else: self._transportes.append(entry)
            self._sync_transp(); self._calc_resumo(); win.destroy()
        bf = ttk.Frame(f); bf.grid(row=5, column=0, columnspan=3, pady=10)
        ModernButton(bf, text="Salvar",   icon=ICONS["save"],   ok=True,     command=salvar).pack(side=tk.LEFT, padx=5)
        ModernButton(bf, text="Cancelar", icon=ICONS["cancel"], danger=True, command=win.destroy).pack(side=tk.LEFT)

    def _edit_transp_sel(self):
        s = self.tree_transp.focus()
        if s:
            try: self._dlg_transporte(idx=int(s))
            except (ValueError,IndexError): pass

    def _del_transp(self):
        s = self.tree_transp.focus()
        if not s: return
        try: self._transportes.pop(int(s))
        except (ValueError,IndexError): return
        self._sync_transp(); self._calc_resumo()

    def _open_transp_anexo(self):
        s = self.tree_transp.focus()
        if not s: return
        try:
            anx = self._transportes[int(s)].get("anexo","")
            if anx and os.path.exists(anx): _abrir_arquivo(anx)
            elif anx: messagebox.showwarning("Aviso", f"Arquivo nao encontrado:\n{anx}")
            else: messagebox.showwarning("Aviso","Nenhum anexo neste registro.")
        except (ValueError,IndexError,OSError) as e:
            messagebox.showerror("Erro", str(e))

    def _sync_transp(self):
        self.tree_transp.delete(*self.tree_transp.get_children())
        for i, t in enumerate(self._transportes):
            anx = "Sim" if t.get("anexo") and os.path.exists(t.get("anexo","")) else (
                  "Link" if t.get("anexo") else "")
            self.tree_transp.insert("", tk.END, iid=str(i), values=(
                t.get("data",""), t.get("tipo",""), t.get("descricao",""),
                f"R$ {t.get('valor',0):.2f}", anx))
        self.lbl_tot_transp.config(
            text=f"Total: R$ {sum(t.get('valor',0) for t in self._transportes):.2f}")

    # ── Materiais CRUD ────────────────────────────────────────────────────────
    def _dlg_material(self, idx=None):
        ex  = self._materiais[idx] if idx is not None else None
        win = tk.Toplevel(self); win.title("Material / Despesa"); win.geometry("430x250"); win.grab_set()
        f   = ttk.Frame(win, padding=15); f.pack(fill=tk.BOTH, expand=True)
        ttk.Label(f, text="Descricao:").grid(row=0, column=0, sticky="w", pady=3)
        e_dc = ttk.Entry(f, width=34); e_dc.insert(0, ex.get("descricao","") if ex else "")
        e_dc.grid(row=0, column=1, columnspan=2, sticky="w")
        ttk.Label(f, text="Quantidade:").grid(row=1, column=0, sticky="w", pady=3)
        e_qt = ttk.Entry(f, width=8); e_qt.insert(0, str(ex.get("qtd",1)) if ex else "1")
        e_qt.grid(row=1, column=1, sticky="w")
        ttk.Label(f, text="Valor unit. (R$):").grid(row=2, column=0, sticky="w", pady=3)
        e_vl = ttk.Entry(f, width=12); e_vl.insert(0, f"{ex.get('valor_unit',0):.2f}" if ex else "0,00")
        e_vl.grid(row=2, column=1, sticky="w")
        ttk.Label(f, text="Nota Fiscal:").grid(row=3, column=0, sticky="w", pady=3)
        e_nf = ttk.Entry(f, width=28); e_nf.insert(0, ex.get("nf","") if ex else "")
        e_nf.grid(row=3, column=1, sticky="w")
        def _browse_nf():
            fp = filedialog.askopenfilename(
                title="Selecionar NF",
                filetypes=[("PDF/Imagem","*.pdf *.png *.jpg *.jpeg"),("Todos","*.*")])
            if fp: e_nf.delete(0,tk.END); e_nf.insert(0,fp)
        ModernButton(f, text="Selecionar", icon=ICONS["attach"], neutral=True,
                     command=_browse_nf).grid(row=3, column=2, padx=4)
        def salvar():
            try:
                qt  = int(e_qt.get().strip())
                vu  = float(e_vl.get().strip().replace(",","."))
            except ValueError:
                qt, vu = 1, 0.0
            entry = {"descricao":e_dc.get().strip(),"qtd":qt,"valor_unit":vu,
                     "total":qt*vu,"nf":e_nf.get().strip()}
            if idx is not None: self._materiais[idx] = entry
            else: self._materiais.append(entry)
            self._sync_mat(); self._calc_resumo(); win.destroy()
        bf = ttk.Frame(f); bf.grid(row=4, column=0, columnspan=3, pady=10)
        ModernButton(bf, text="Salvar",   icon=ICONS["save"],   ok=True,     command=salvar).pack(side=tk.LEFT, padx=5)
        ModernButton(bf, text="Cancelar", icon=ICONS["cancel"], danger=True, command=win.destroy).pack(side=tk.LEFT)

    def _edit_mat_sel(self):
        s = self.tree_mat.focus()
        if s:
            try: self._dlg_material(idx=int(s))
            except (ValueError,IndexError): pass

    def _del_mat(self):
        s = self.tree_mat.focus()
        if not s: return
        try: self._materiais.pop(int(s))
        except (ValueError,IndexError): return
        self._sync_mat(); self._calc_resumo()

    def _open_mat_nf(self):
        s = self.tree_mat.focus()
        if not s: return
        try:
            nf = self._materiais[int(s)].get("nf","")
            if nf and os.path.exists(nf): _abrir_arquivo(nf)
            elif nf: messagebox.showwarning("Aviso", f"Arquivo nao encontrado:\n{nf}")
            else: messagebox.showwarning("Aviso","Nenhuma NF anexada.")
        except (ValueError,IndexError,OSError) as e:
            messagebox.showerror("Erro", str(e))

    def _sync_mat(self):
        self.tree_mat.delete(*self.tree_mat.get_children())
        for i, m in enumerate(self._materiais):
            nf_ok = "Sim" if m.get("nf") and os.path.exists(m.get("nf","")) else (
                    "Link" if m.get("nf") else "")
            self.tree_mat.insert("", tk.END, iid=str(i), values=(
                m.get("descricao",""), m.get("qtd",1),
                f"R$ {m.get('valor_unit',0):.2f}",
                f"R$ {m.get('total',0):.2f}", nf_ok))
        self.lbl_tot_mat.config(
            text=f"Total: R$ {sum(m.get('total',0) for m in self._materiais):.2f}")

    # ── Links panel ───────────────────────────────────────────────────────────
    def _refresh_links(self):
        for w in self._links_ctn.winfo_children(): w.destroy()
        links = []
        for i, h in enumerate(self.hotel_entries):
            nome = h["nome"].get().strip() or f"Hotel {i+1}"
            lk   = h["link"].get().strip()
            if lk and lk not in ("https://","http://") and lk.startswith("http"):
                links.append(("Hotel", nome, lk))
        for p in self._passagens:
            lk = p.get("link","")
            if lk and lk.startswith("http"):
                links.append(("Passagem", f"{p.get('tipo','')} {p.get('trecho','')}".strip(), lk))
        if not links:
            ttk.Label(self._links_ctn,
                text="(preencha hospedagem ou passagens para ver os links aqui)",
                font=("Arial",8), foreground="gray").pack(anchor="w")
            return
        for cat, nome, lk in links:
            row_f = ttk.Frame(self._links_ctn); row_f.pack(fill=tk.X, pady=2)
            tk.Label(row_f, text=f"[{cat}]", font=("Arial",8,"bold"),
                     fg="#2563EB", width=10, anchor="w").pack(side=tk.LEFT)
            ttk.Label(row_f, text=nome, font=("Arial",9,"bold"),
                      width=28, anchor="w").pack(side=tk.LEFT)
            ModernButton(row_f, text="Abrir", icon=ICONS["open"], neutral=True,
                         command=lambda u=lk: webbrowser.open(u)).pack(side=tk.LEFT, padx=(0,8))
            short = (lk[:55]+"...") if len(lk)>55 else lk
            ttk.Label(row_f, text=short, font=("Courier",8),
                      foreground="gray").pack(side=tk.LEFT)

    # ── Gerar Canhoto ─────────────────────────────────────────────────────────
    def _gerar_canhoto(self):
        nome  = self.e_nome.get().strip()
        dest  = self.e_destino.get().strip()
        ida   = self.e_ida.get().strip()
        volta = self.e_volta.get().strip()
        h_ida = self.e_hora_ida.get().strip()
        h_vta = self.e_hora_volta.get().strip()
        pix   = self.e_pix.get().strip()
        obs   = self.txt_obs.get("1.0", tk.END).strip()
        if h_ida  == "HH:MM": h_ida  = ""
        if h_vta  == "HH:MM": h_vta  = ""

        if not all([nome, dest, ida, volta, pix]):
            messagebox.showwarning("Aviso","Preencha: Nome, Destino, Datas e PIX."); return
        dias = self._days(ida, volta)
        if dias == 0:
            messagebox.showerror("Erro","Datas invalidas! Use DD/MM/AAAA."); return
        for h, lb in [(h_ida,"partida"),(h_vta,"retorno")]:
            if h:
                try: datetime.strptime(h, "%H:%M")
                except ValueError:
                    messagebox.showerror("Erro", f"Hora de {lb} invalida (HH:MM)."); return

        vd         = self.cfg["valor_diaria"]
        hotel      = self._get_hotel()
        val_d      = dias * vd
        val_p      = sum(p.get("valor",0) for p in self._passagens)
        val_t      = sum(t.get("valor",0) for t in self._transportes)
        val_m      = sum(m.get("total",0) for m in self._materiais)
        val_h      = hotel["valor_noite"] * dias if hotel["selecionado"] == "Sim" else 0
        val_total  = val_d + val_p + val_t + val_h + val_m

        part_str = ida   + (f" as {h_ida}" if h_ida else "")
        ret_str  = volta + (f" as {h_vta}" if h_vta else "")

        # Bloco hospedagem — todas as 3 opcoes e links
        h_bloco = f"Utilizou Hotel  : {hotel['selecionado']}"
        if hotel["selecionado"] == "Sim":
            h_bloco += f"\nHotel Escolhido : {hotel['nome']}"
            if hotel["valor_noite"] > 0:
                h_bloco += f" (R$ {hotel['valor_noite']:.2f}/noite)"
            if hotel.get("link","") not in ("","https://"):
                h_bloco += f"\nLink do Hotel   : {hotel['link']}"
        if hotel.get("todas_opcoes"):
            h_bloco += "\nOpcoes pesquisadas:"
            for op in hotel["todas_opcoes"]:
                sel = " [ESCOLHIDA]" if op["selecionada"] else ""
                h_bloco += f"\n  {op['opcao']}. {op['nome']} — R$ {op['valor_noite']:.2f}/noite{sel}"
                if op.get("link","") not in ("","https://"):
                    h_bloco += f"\n     Link: {op['link']}"

        # Bloco passagens
        p_bloco = ""
        if self._passagens:
            p_bloco = f"\nPASSAGENS ({len(self._passagens)} bilhete(s)):\n"
            for p in self._passagens:
                dh  = p.get("data","") + (f" {p['hora']}" if p.get("hora") else "")
                loc = f" | Loc: {p['localizador']}" if p.get("localizador") else ""
                p_bloco += f"  [{p.get('tipo','')}] {p.get('trecho','')} — {dh} — R$ {p.get('valor',0):.2f}{loc}\n"
                if p.get("link","") not in ("","https://"):
                    p_bloco += f"  Link: {p['link']}\n"
            p_bloco += f"  Subtotal: R$ {val_p:.2f}\n"

        # Bloco transporte
        t_bloco = ""
        if self._transportes:
            t_bloco = f"\nTRANSPORTE LOCAL ({len(self._transportes)} lancamento(s)):\n"
            for t in self._transportes:
                t_bloco += f"  {t.get('data','')} [{t.get('tipo','')}] {t.get('descricao','')} — R$ {t.get('valor',0):.2f}"
                if t.get("anexo"): t_bloco += " [anexo]"
                t_bloco += "\n"
            t_bloco += f"  Subtotal: R$ {val_t:.2f}\n"

        # Bloco materiais
        m_bloco = ""
        if self._materiais:
            m_bloco = f"\nMATERIAIS / DESPESAS ({len(self._materiais)} item(ns)):\n"
            for m in self._materiais:
                m_bloco += f"  {m.get('descricao','')} x{m.get('qtd',1)} — R$ {m.get('total',0):.2f}"
                if m.get("nf"): m_bloco += " [NF]"
                m_bloco += "\n"
            m_bloco += f"  Subtotal: R$ {val_m:.2f}\n"

        canhoto = (
            f"\n{'='*60}\n"
            f"              CANHOTO DE VIAGEM — \n"
            f"{'='*60}\n"
            f"Viajante        : {nome}\n"
            f"Destino         : {dest}\n"
            f"Partida         : {part_str}\n"
            f"Retorno         : {ret_str}\n"
            f"{h_bloco}\n"
            f"Dias            : {dias} dia(s)\n"
            f"Valor diarias   : R$ {val_d:.2f}  (R$ {vd:.2f} x {dias})\n"
            f"PIX             : {pix}\n"
            + (f"Observacoes     : {obs}\n" if obs else "")
            + p_bloco + t_bloco + m_bloco
            + f"\n{'─'*60}\n"
            f"  Diarias        : R$ {val_d:.2f}\n"
            f"  Passagens      : R$ {val_p:.2f}\n"
            f"  Transporte     : R$ {val_t:.2f}\n"
            f"  Hospedagem     : R$ {val_h:.2f}\n"
            f"  Materiais      : R$ {val_m:.2f}\n"
            f"  TOTAL ESTIMADO : R$ {val_total:.2f}\n"
            f"{'─'*60}\n"
            f"Emitido em      : {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
            f"{'='*60}\n"
        )

        reg_id = str(uuid.uuid4())[:8]
        registro = {
            "id": reg_id, "nome": nome, "destino": dest,
            "ida": ida, "hora_ida": h_ida,
            "volta": volta, "hora_volta": h_vta,
            "hotel": hotel,
            "passagens":   list(self._passagens),
            "transportes": list(self._transportes),
            "materiais":   list(self._materiais),
            "dias": dias,
            "valor":          val_d,
            "val_passagens":  val_p,
            "val_transporte": val_t,
            "val_hotel":      val_h,
            "val_materiais":  val_m,
            "val_total":      val_total,
            "pix": pix, "obs": obs,
            "data_registro": datetime.now().strftime("%d/%m/%Y %H:%M"),
        }
        dados = self._load_data()
        dados.append(registro)
        self._save_data(dados)
        self._refresh_status()
        self._janela_canhoto(canhoto, pix)

        if self.cfg.get("auto_email", False):
            self.after(400, lambda: self._enviar_auto(canhoto=canhoto))

    def _janela_canhoto(self, canhoto, pix):
        win = tk.Toplevel(self); win.title("Canhoto"); win.geometry("660x540"); win.grab_set()
        tk.Label(win, text="Viagem registrada com sucesso!",
                 font=("Arial",11,"bold"), fg="#16A34A").pack(pady=(10,0))
        txt = tk.Text(win, height=22, width=78, font=("Courier",10))
        txt.insert("1.0", canhoto); txt.config(state=tk.DISABLED)
        txt.pack(padx=10, pady=8)
        bf = ttk.Frame(win); bf.pack(pady=6)
        def _cp_pix():
            try: pyperclip.copy(pix); messagebox.showinfo("OK","PIX copiado!", parent=win)
            except Exception as e: messagebox.showerror("Erro",str(e),parent=win)
        def _cp_can():
            try: pyperclip.copy(canhoto); messagebox.showinfo("OK","Canhoto copiado!", parent=win)
            except Exception as e: messagebox.showerror("Erro",str(e),parent=win)
        ModernButton(bf, text="Copiar PIX",     icon=ICONS["pix"],    ok=True,      command=_cp_pix).pack(side=tk.LEFT,padx=4)
        ModernButton(bf, text="Copiar Canhoto",icon=ICONS["copy"],               command=_cp_can).pack(side=tk.LEFT,padx=4)
        ModernButton(bf, text="Salvar TXT/PDF",icon=ICONS["backup"], neutral=True,command=lambda: self._save_file(canhoto)).pack(side=tk.LEFT,padx=4)
        ModernButton(bf, text="Enviar E-mail", icon=ICONS["email"],  ok=True,     command=lambda: self._dlg_email(canhoto,win)).pack(side=tk.LEFT,padx=4)
        ModernButton(bf, text="Fechar",        icon=ICONS["cancel"], danger=True, command=win.destroy).pack(side=tk.LEFT,padx=4)

    def _save_file(self, canhoto):
        tipos = [("Texto","*.txt")]
        ext   = ".txt"
        if _REPORTLAB: tipos.insert(0,("PDF","*.pdf")); ext=".pdf"
        arq = filedialog.asksaveasfilename(
            defaultextension=ext, filetypes=tipos,
            initialfile=f"canhoto_{datetime.now().strftime('%Y%m%d_%H%M')}")
        if not arq: return
        if arq.lower().endswith(".pdf") and _REPORTLAB:
            try:
                c = rl_canvas.Canvas(arq, pagesize=A4)
                c.setFont("Courier",11)
                _, h = A4; y = h-60
                for ln in canhoto.splitlines():
                    c.drawString(50, y, ln.encode("latin-1","replace").decode("latin-1"))
                    y -= 16
                    if y < 50: c.showPage(); c.setFont("Courier",11); y = h-60
                c.save()
                messagebox.showinfo("Salvo", f"PDF:\n{arq}")
            except Exception as e: messagebox.showerror("Erro",str(e))
        else:
            try:
                with open(arq,"w",encoding="utf-8") as fp: fp.write(canhoto)
                messagebox.showinfo("Salvo", f"TXT:\n{arq}")
            except IOError as e: messagebox.showerror("Erro",str(e))

    # ── E-mail ────────────────────────────────────────────────────────────────
    def _dlg_email(self, canhoto, parent=None):
        cfg = self.cfg
        win = tk.Toplevel(parent or self)
        win.title("Enviar por E-mail"); win.geometry("480x300"); win.grab_set()
        f = ttk.Frame(win, padding=15); f.pack(fill=tk.BOTH, expand=True)
        ttk.Label(f, text="Enviar Canhoto",
                  font=("Arial",12,"bold")).grid(row=0,column=0,columnspan=2,pady=(0,10))
        campos = [
            ("De:",            cfg.get("email_remetente",""),    False),
            ("Senha:",         self._get_email_senha(),           True),
            ("Para:",          cfg.get("email_destinatario",""),  False),
            ("Assunto:",       "Viagem agendada",                 False),
            ("SMTP:",          cfg.get("email_smtp","smtp.gmail.com"), False),
            ("Porta:",         str(cfg.get("email_porta",587)),   False),
        ]
        ents = {}
        for i,(lbl,val,sec) in enumerate(campos):
            ttk.Label(f,text=lbl).grid(row=i+1,column=0,sticky="w",pady=3,padx=(0,8))
            e = ttk.Entry(f,width=34,show="*" if sec else ""); e.insert(0,val)
            e.grid(row=i+1,column=1,sticky="w",pady=3); ents[lbl]=e
        def _enviar():
            rem  = ents["De:"].get().strip()
            pwd  = ents["Senha:"].get().strip()
            dest = ents["Para:"].get().strip()
            subj = ents["Assunto:"].get().strip() or "Viagem agendada"
            srv  = ents["SMTP:"].get().strip() or "smtp.gmail.com"
            try: porta=int(ents["Porta:"].get().strip())
            except ValueError: porta=587
            if not all([rem,pwd,dest]):
                messagebox.showwarning("Aviso","Preencha De, Senha e Para.",parent=win); return
            win.destroy()
            threading.Thread(target=self._smtp_send,
                             args=(rem, pwd, dest, subj, srv, porta, canhoto),
                             daemon=True).start()
        bf = ttk.Frame(f)
        bf.grid(row=len(campos)+1, column=0, columnspan=2, pady=12)
        ModernButton(bf, text="Enviar",   icon=ICONS["email"],  ok=True,     command=_enviar).pack(side=tk.LEFT,padx=5)
        ModernButton(bf, text="Cancelar", icon=ICONS["cancel"], danger=True, command=win.destroy).pack(side=tk.LEFT)

    def _enviar_auto(self, canhoto="", assunto="Viagem agendada"):
        cfg  = self.cfg
        rem  = cfg.get("email_remetente","").strip()
        pwd  = self._get_email_senha().strip()
        dest = cfg.get("email_destinatario","").strip()
        srv  = cfg.get("email_smtp","smtp.gmail.com")
        try: porta=int(cfg.get("email_porta",587))
        except (ValueError,TypeError): porta=587
        if not all([rem,pwd,dest]):
            messagebox.showwarning("E-mail nao configurado",
                "Preencha remetente, senha e destinatario em Configuracoes > E-mail.")
            return
        self.status_var.set("Enviando e-mail...")
        self.update_idletasks()
        threading.Thread(target=self._smtp_send,
                         args=(rem, pwd, dest, assunto, srv, porta, canhoto),
                         daemon=True).start()

    def _smtp_send(self, rem, pwd, dest, assunto, srv, porta, canhoto):
        """Executa em thread separada — nunca chame widgets diretamente aqui."""
        try:
            msg = MIMEMultipart("alternative")
            msg["Subject"] = assunto; msg["From"] = rem; msg["To"] = dest
            now_str    = datetime.now().strftime("%d/%m/%Y %H:%M")
            corpo_txt  = f"Canhoto gerado em {now_str}.\n\n{canhoto}"
            corpo_html = (
                "<html><body>"
                f"<p><b>Canhoto de viagem</b> — {html.escape(now_str)}</p>"
                f'<pre style="font-family:Courier,monospace;font-size:13px;">'
                f"{html.escape(canhoto)}</pre>"
                "</body></html>"
            )
            msg.attach(MIMEText(corpo_txt, "plain", "utf-8"))
            msg.attach(MIMEText(corpo_html, "html", "utf-8"))
            with smtplib.SMTP(srv, porta, timeout=15) as s:
                s.ehlo(); s.starttls(); s.ehlo()
                s.login(rem, pwd)
                s.sendmail(rem, dest, msg.as_string())
            def _ok():
                messagebox.showinfo("Enviado", f"E-mail enviado para:\n{dest}")
                self._refresh_status()
            self.after(0, _ok)
        except smtplib.SMTPAuthenticationError:
            self.after(0, lambda: (
                messagebox.showerror("Autenticacao falhou",
                    "Senha incorreta.\nGmail/Workspace: use Senha de App em "
                    "myaccount.google.com > Senhas de app."),
                self._refresh_status()
            ))
        except Exception as e:
            msg_err = str(e)
            self.after(0, lambda: (
                messagebox.showerror("Erro ao enviar", msg_err),
                self._refresh_status()
            ))

    # ── Limpar ────────────────────────────────────────────────────────────────
    def _limpar(self):
        if not messagebox.askyesno("Limpar","Apagar todos os campos?"): return
        for e in (self.e_nome, self.e_destino, self.e_ida, self.e_volta, self.e_pix):
            e.delete(0, tk.END)
        for e, ph in [(self.e_hora_ida,"HH:MM"),(self.e_hora_volta,"HH:MM")]:
            e.delete(0, tk.END); e.insert(0, ph)
        self.txt_obs.delete("1.0", tk.END)
        self.var_hotel.set(0)
        for i, h in enumerate(self.hotel_entries):
            h["nome"].delete(0, tk.END);  h["nome"].insert(0, f"Opcao {i+1}")
            h["link"].delete(0, tk.END);  h["link"].insert(0, "https://")
            h["valor"].delete(0, tk.END); h["valor"].insert(0, "0,00")
        self.lbl_dias.config(text="—"); self.lbl_val_diarias.config(text="R$ 0,00")
        for lb in (self.lbl_ida_ok, self.lbl_volta_ok, self.lbl_pix_ok):
            lb.config(text="")
        self._passagens.clear(); self._transportes.clear(); self._materiais.clear()
        self._sync_pass(); self._sync_transp(); self._sync_mat()
        self._calc_resumo(); self._refresh_links()

    # ── Historico ─────────────────────────────────────────────────────────────
    def _load_hist(self):
        dados = self._load_data()
        anos  = ["Todos"] + sorted(
            {d.get("ida","")[-4:] for d in dados if len(d.get("ida",""))==10}, reverse=True)
        self.cmb_ano["values"] = anos
        if self.cmb_ano.get() not in anos: self.cmb_ano.current(0)
        self.tree_hist.delete(*self.tree_hist.get_children())
        self._fill_hist(dados)

    def _filter_hist(self):
        tnome = self.e_busca_nome.get().lower()
        tdest = self.e_busca_dest.get().lower()
        mes   = self.cmb_mes.current()
        ano   = self.cmb_ano.get()
        res   = []
        for d in self._load_data():
            ok_n = tnome in d.get("nome","").lower()
            ok_d = tdest in d.get("destino","").lower()
            ok_m = ok_a = True
            ida  = d.get("ida","")
            if len(ida)==10:
                p = ida.split("/")
                if mes>0:          ok_m = int(p[1])==mes
                if ano!="Todos":   ok_a = p[2]==ano
            if ok_n and ok_d and ok_m and ok_a:
                res.append(d)
        self.tree_hist.delete(*self.tree_hist.get_children())
        self._fill_hist(res)

    def _fill_hist(self, dados):
        total = 0
        for d in dados:
            h  = d.get("hotel", {})
            if isinstance(h, dict) and h.get("selecionado") == "Sim":
                hn = h.get("nome", "") or "Sim"
            else:
                hn = "Nao"
            partida = d.get("ida","") + (f" {d['hora_ida']}"   if d.get("hora_ida")   else "")
            retorno = d.get("volta","")+(f" {d['hora_volta']}" if d.get("hora_volta") else "")
            vt      = d.get("val_total", d.get("valor",0))
            self.tree_hist.insert("", tk.END, iid=d.get("id",""), values=(
                d.get("nome",""), d.get("destino",""),
                partida, retorno, hn,
                len(d.get("passagens",[])) or "",
                len(d.get("materiais",[])) or "",
                d.get("dias",""),
                f"R$ {vt:.2f}"))
            total += vt
        self.lbl_hist_total.config(
            text=f"Total: R$ {total:.2f}  |  {len(dados)} registro(s)")

    def _open_edit_dlg(self, _e=None):
        sel = self.tree_hist.focus()
        if not sel: return
        dados = self._load_data()
        reg   = next((d for d in dados if d.get("id")==sel), None)
        if not reg: return
        self._dlg_editar(reg)

    def _dlg_editar(self, reg):
        win = tk.Toplevel(self); win.title("Editar Registro"); win.geometry("560x560"); win.grab_set()
        nb  = ttk.Notebook(win); nb.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        # Tab dados
        f1 = ttk.Frame(nb, padding=12); nb.add(f1, text="  Dados  ")
        campos = [("Nome:","nome"),("Destino:","destino"),("Data Ida:","ida"),
                  ("Hora Ida:","hora_ida"),("Data Volta:","volta"),("Hora Volta:","hora_volta"),("PIX:","pix")]
        ents = {}
        for i,(lb,k) in enumerate(campos):
            ttk.Label(f1,text=lb,font=("Arial",9,"bold")).grid(row=i,column=0,sticky="w",pady=3,padx=(0,10))
            e = ttk.Entry(f1,width=36); e.insert(0,reg.get(k,"")); e.grid(row=i,column=1,sticky="w",pady=3)
            ents[k]=e
        ttk.Label(f1,text="Obs:",font=("Arial",9,"bold")).grid(row=len(campos),column=0,sticky="nw",pady=3)
        t_obs = tk.Text(f1,height=3,width=36,font=("Arial",9))
        t_obs.insert("1.0",reg.get("obs","")); t_obs.grid(row=len(campos),column=1,sticky="w")

        # Tab passagens
        f2 = ttk.Frame(nb, padding=12); nb.add(f2, text="  Passagens  ")
        if reg.get("passagens"):
            cols=("Tipo","Trecho","Data","Valor","Loc.")
            tr=ttk.Treeview(f2,columns=cols,show="headings",height=6)
            for c in cols: tr.heading(c,text=c); tr.column(c,width=90)
            for p in reg["passagens"]:
                tr.insert("",tk.END,values=(p.get("tipo",""),p.get("trecho",""),
                    p.get("data",""),f"R$ {p.get('valor',0):.2f}",p.get("localizador","")))
            tr.pack(fill=tk.BOTH,expand=True)
        else:
            ttk.Label(f2,text="Nenhuma passagem.",foreground="gray").pack(pady=20)

        # Tab transportes (editavel)
        f3 = ttk.Frame(nb, padding=12); nb.add(f3, text="  Transporte  ")
        trans_list = list(reg.get("transportes",[]))
        cols_t = ("Data","Tipo","Descricao","Valor","Anexo")
        tr_t   = ttk.Treeview(f3,columns=cols_t,show="headings",height=6)
        for c in cols_t: tr_t.heading(c,text=c); tr_t.column(c,width=90)
        def _sync_tr_t():
            tr_t.delete(*tr_t.get_children())
            for i,t in enumerate(trans_list):
                anx = "Sim" if t.get("anexo") and os.path.exists(t.get("anexo","")) else ("" if not t.get("anexo") else "Link")
                tr_t.insert("",tk.END,iid=str(i),values=(t.get("data",""),t.get("tipo",""),
                    t.get("descricao",""),f"R$ {t.get('valor',0):.2f}",anx))
        _sync_tr_t()
        tr_t.pack(fill=tk.BOTH,expand=True)
        bf3 = ttk.Frame(f3); bf3.pack(fill=tk.X, pady=4)
        def _add_t():
            def _save_new(entry):
                trans_list.append(entry); _sync_tr_t()
            self._dlg_transporte_inline(_save_new, bf3)
        def _del_t():
            s = tr_t.focus()
            if s:
                try: trans_list.pop(int(s)); _sync_tr_t()
                except (ValueError,IndexError): pass
        ModernButton(bf3, text="Adicionar", icon=ICONS["add"],    command=_add_t).pack(side=tk.LEFT,padx=(0,4))
        ModernButton(bf3, text="Remover",   icon=ICONS["remove"], command=_del_t, danger=True).pack(side=tk.LEFT)

        # Tab materiais (editavel)
        f4 = ttk.Frame(nb, padding=12); nb.add(f4, text="  Materiais  ")
        mat_list = list(reg.get("materiais",[]))
        cols_m = ("Descricao","Qtd","Valor Unit","Total","NF")
        tr_m   = ttk.Treeview(f4,columns=cols_m,show="headings",height=6)
        for c in cols_m: tr_m.heading(c,text=c); tr_m.column(c,width=85)
        def _sync_tr_m():
            tr_m.delete(*tr_m.get_children())
            for i,m in enumerate(mat_list):
                nf = "Sim" if m.get("nf") and os.path.exists(m.get("nf","")) else ("" if not m.get("nf") else "Link")
                tr_m.insert("",tk.END,iid=str(i),values=(m.get("descricao",""),m.get("qtd",1),
                    f"R$ {m.get('valor_unit',0):.2f}",f"R$ {m.get('total',0):.2f}",nf))
        _sync_tr_m()
        tr_m.pack(fill=tk.BOTH,expand=True)
        bf4 = ttk.Frame(f4); bf4.pack(fill=tk.X,pady=4)
        def _add_m():
            def _save_new(entry):
                mat_list.append(entry); _sync_tr_m()
            self._dlg_material_inline(_save_new, bf4)
        def _del_m():
            s = tr_m.focus()
            if s:
                try: mat_list.pop(int(s)); _sync_tr_m()
                except (ValueError,IndexError): pass
        ModernButton(bf4, text="Adicionar", icon=ICONS["add"],    command=_add_m).pack(side=tk.LEFT,padx=(0,4))
        ModernButton(bf4, text="Remover",   icon=ICONS["remove"], command=_del_m, danger=True).pack(side=tk.LEFT)

        def _salvar():
            dados = self._load_data()
            for j, d in enumerate(dados):
                if d.get("id") == reg.get("id"):
                    for k, e in ents.items():
                        dados[j][k] = e.get().strip()
                    dados[j]["obs"]         = t_obs.get("1.0", tk.END).strip()
                    dados[j]["transportes"] = trans_list
                    dados[j]["materiais"]   = mat_list
                    d2 = self._days(dados[j]["ida"], dados[j]["volta"])
                    if d2 > 0:
                        dados[j]["dias"]  = d2
                        dados[j]["valor"] = d2 * self.cfg["valor_diaria"]
                    # recalcula todos os subtotais derivados
                    dados[j]["val_passagens"]  = sum(p.get("valor", 0) for p in dados[j].get("passagens", []))
                    dados[j]["val_transporte"] = sum(t.get("valor", 0) for t in trans_list)
                    dados[j]["val_materiais"]  = sum(m.get("total", 0) for m in mat_list)
                    h = dados[j].get("hotel", {})
                    val_h = h.get("valor_noite", 0) * d2 if isinstance(h, dict) and h.get("selecionado") == "Sim" and d2 > 0 else 0
                    dados[j]["val_hotel"] = val_h
                    dados[j]["val_total"] = (dados[j]["valor"] + dados[j]["val_passagens"] +
                                             dados[j]["val_transporte"] + val_h + dados[j]["val_materiais"])
                    break
            self._save_data(dados); self._load_hist(); self._refresh_status()
            messagebox.showinfo("Salvo", "Registro atualizado!", parent=win); win.destroy()

        bf = ttk.Frame(win); bf.pack(pady=8)
        ModernButton(bf, text="Salvar alteracoes", icon=ICONS["save"],   ok=True,     command=_salvar).pack(side=tk.LEFT,padx=5)
        ModernButton(bf, text="Cancelar",          icon=ICONS["cancel"], danger=True, command=win.destroy).pack(side=tk.LEFT)

    def _dlg_transporte_inline(self, callback, parent):
        win = tk.Toplevel(parent.winfo_toplevel())
        win.title("Novo Transporte"); win.geometry("400x200"); win.grab_set()
        f = ttk.Frame(win,padding=12); f.pack(fill=tk.BOTH,expand=True)
        fields = [("Data:",""),("Descricao:",""),("Valor (R$):","0,00")]
        ttk.Label(f,text="Tipo:").grid(row=0,column=0,sticky="w",pady=3)
        cmb = ttk.Combobox(f,values=TIPOS_TRANSPORTE,state="readonly",width=16)
        cmb.set("Uber"); cmb.grid(row=0,column=1,sticky="w")
        ents = {}
        for i,(lb,val) in enumerate(fields):
            ttk.Label(f,text=lb).grid(row=i+1,column=0,sticky="w",pady=3)
            e=ttk.Entry(f,width=28); e.insert(0,val); e.grid(row=i+1,column=1,sticky="w"); ents[lb]=e
        def _save():
            try: v=float(ents["Valor (R$):"].get().strip().replace(",","."))
            except ValueError: v=0.0
            callback({"data":ents["Data:"].get().strip(),"tipo":cmb.get(),
                      "descricao":ents["Descricao:"].get().strip(),"valor":v,"anexo":""})
            win.destroy()
        bf=ttk.Frame(f); bf.grid(row=4,column=0,columnspan=2,pady=8)
        ModernButton(bf, text="Salvar",   icon=ICONS["save"],   ok=True,     command=_save).pack(side=tk.LEFT,padx=4)
        ModernButton(bf, text="Cancelar", icon=ICONS["cancel"], danger=True, command=win.destroy).pack(side=tk.LEFT)

    def _dlg_material_inline(self, callback, parent):
        win = tk.Toplevel(parent.winfo_toplevel())
        win.title("Novo Material"); win.geometry("400x200"); win.grab_set()
        f = ttk.Frame(win,padding=12); f.pack(fill=tk.BOTH,expand=True)
        fields = [("Descricao:",""),("Quantidade:","1"),("Valor unit. (R$):","0,00")]
        ents={}
        for i,(lb,val) in enumerate(fields):
            ttk.Label(f,text=lb).grid(row=i,column=0,sticky="w",pady=3)
            e=ttk.Entry(f,width=28); e.insert(0,val); e.grid(row=i,column=1,sticky="w"); ents[lb]=e
        def _save():
            try: qt=int(ents["Quantidade:"].get().strip()); vu=float(ents["Valor unit. (R$):"].get().strip().replace(",","."))
            except ValueError: qt,vu=1,0.0
            callback({"descricao":ents["Descricao:"].get().strip(),"qtd":qt,
                      "valor_unit":vu,"total":qt*vu,"nf":""})
            win.destroy()
        bf=ttk.Frame(f); bf.grid(row=3,column=0,columnspan=2,pady=8)
        ModernButton(bf, text="Salvar",   icon=ICONS["save"],   ok=True,     command=_save).pack(side=tk.LEFT,padx=4)
        ModernButton(bf, text="Cancelar", icon=ICONS["cancel"], danger=True, command=win.destroy).pack(side=tk.LEFT)

    def _del_registro(self):
        sel = self.tree_hist.focus()
        if not sel: messagebox.showwarning("Aviso","Selecione um registro."); return
        if not messagebox.askyesno("Confirmar","Excluir permanentemente?"): return
        dados = [d for d in self._load_data() if d.get("id")!=sel]
        self._save_data(dados); self._load_hist(); self._refresh_status()

    # ── Excel (3 abas) ────────────────────────────────────────────────────────
    def _exportar_excel(self):
        dados = self._load_data()
        if not dados: messagebox.showwarning("Aviso","Nenhum dado."); return
        arq = filedialog.asksaveasfilename(
            defaultextension=".xlsx", filetypes=[("Excel","*.xlsx")],
            initialfile=f"viagens_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")
        if not arq: return
        wb   = Workbook()
        fill = PatternFill(start_color="1E3A5F",end_color="1E3A5F",fill_type="solid")
        fnt  = Font(color="FFFFFF",bold=True)
        def _hdr(ws, cols):
            ws.append(cols)
            for c in ws[1]: c.fill=fill; c.font=fnt; c.alignment=Alignment(horizontal="center")
        # Viagens
        ws1 = wb.active; ws1.title="Viagens"
        _hdr(ws1,["Nome","Destino","Ida","Hora Ida","Volta","Hora Volta",
                  "Hotel","Nome Hotel","Link Hotel","Val.Hotel/Noite","Dias",
                  "Diarias","Passagens","Transporte","Materiais","TOTAL","PIX","Obs","Registro"])
        total_g = 0
        for d in dados:
            h  = d.get("hotel",{})
            vt = d.get("val_total",d.get("valor",0))
            ws1.append([d.get("nome",""),d.get("destino",""),
                        d.get("ida",""),d.get("hora_ida",""),
                        d.get("volta",""),d.get("hora_volta",""),
                        h.get("selecionado","") if isinstance(h,dict) else "",
                        h.get("nome","")        if isinstance(h,dict) else "",
                        h.get("link","")        if isinstance(h,dict) else "",
                        h.get("valor_noite",0)  if isinstance(h,dict) else 0,
                        d.get("dias",0),
                        d.get("valor",0),d.get("val_passagens",0),
                        d.get("val_transporte",0),d.get("val_materiais",0),vt,
                        d.get("pix",""),d.get("obs",""),d.get("data_registro","")]); total_g+=vt
        ul=len(dados)+2; ws1[f"A{ul}"]="TOTAL"; ws1[f"P{ul}"]=total_g
        ws1[f"P{ul}"].font=Font(bold=True)
        for col,w in zip("ABCDEFGHIJKLMNOPQRS",
                         [22,20,12,9,12,9,8,22,35,14,6,12,12,12,12,14,22,25,18]):
            ws1.column_dimensions[col].width=w
        # Passagens
        ws2=wb.create_sheet("Passagens")
        _hdr(ws2,["Viajante","Destino","Tipo","Trecho","Data","Hora","Valor","Localizador","Link","Registro"])
        for d in dados:
            for p in d.get("passagens",[]):
                ws2.append([d.get("nome",""),d.get("destino",""),
                    p.get("tipo",""),p.get("trecho",""),p.get("data",""),p.get("hora",""),
                    p.get("valor",0),p.get("localizador",""),p.get("link",""),d.get("data_registro","")])
        for col,w in zip("ABCDEFGHIJ",[22,20,10,25,12,8,12,14,40,18]): ws2.column_dimensions[col].width=w
        # Transportes
        ws3=wb.create_sheet("Transporte")
        _hdr(ws3,["Viajante","Destino","Data","Tipo","Descricao","Valor","Anexo","Registro"])
        for d in dados:
            for t in d.get("transportes",[]):
                ws3.append([d.get("nome",""),d.get("destino",""),
                    t.get("data",""),t.get("tipo",""),t.get("descricao",""),
                    t.get("valor",0),t.get("anexo",""),d.get("data_registro","")])
        for col,w in zip("ABCDEFGH",[22,20,12,10,35,12,30,18]): ws3.column_dimensions[col].width=w
        # Materiais
        ws4=wb.create_sheet("Materiais")
        _hdr(ws4,["Viajante","Destino","Descricao","Qtd","Valor Unit","Total","NF","Registro"])
        for d in dados:
            for m in d.get("materiais",[]):
                ws4.append([d.get("nome",""),d.get("destino",""),
                    m.get("descricao",""),m.get("qtd",1),
                    m.get("valor_unit",0),m.get("total",0),m.get("nf",""),d.get("data_registro","")])
        for col,w in zip("ABCDEFGH",[22,20,30,6,12,12,35,18]): ws4.column_dimensions[col].width=w

        wb.save(arq)
        messagebox.showinfo("Exportado",
            f"Excel com 4 abas (Viagens, Passagens, Transporte, Materiais):\n{arq}")

    # ── Persistencia ──────────────────────────────────────────────────────────
    def _load_data(self):
        if os.path.exists(DATA_FILE):
            try:
                with open(DATA_FILE,"r",encoding="utf-8") as f:
                    dados = json.load(f)
                changed = False
                for d in dados:
                    if "id" not in d: d["id"]=str(uuid.uuid4())[:8]; changed=True
                if changed: self._save_data(dados)
                return dados
            except (json.JSONDecodeError,IOError):
                messagebox.showwarning("Aviso","Arquivo corrompido. Iniciando do zero.")
        return []

    def _save_data(self, dados):
        try:
            with open(DATA_FILE,"w",encoding="utf-8") as f:
                json.dump(dados,f,ensure_ascii=False,indent=4)
        except IOError as e:
            messagebox.showerror("Erro",f"Nao foi possivel salvar:\n{e}")

    def _refresh_status(self):
        dados = self._load_data()
        total = sum(d.get("val_total",d.get("valor",0)) for d in dados)
        self.status_var.set(
            f"  {len(dados)} viagem(ns) registrada(s)  |  Total acumulado: R$ {total:.2f}")


# ── Ponto de entrada ──────────────────────────────────────────────────────────
if __name__ == "__main__":
    AppViagens().mainloop()