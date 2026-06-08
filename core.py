# -*- coding: utf-8 -*-
"""core.py — Lógica de negócio compartilhada entre Tkinter e Streamlit."""
from __future__ import annotations
import json, os, re, shutil, uuid, hashlib, html, smtplib
from datetime import datetime
from pathlib import Path
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    _EXCEL = True
except ImportError:
    _EXCEL = False

try:
    from cryptography.fernet import Fernet
    import base64
    _CRYPTO = True
except ImportError:
    _CRYPTO = False

# ── Caminhos ──────────────────────────────────────────────────────────────────
BASE_DIR    = Path(__file__).parent
DATA_FILE   = str(BASE_DIR / "viagens.json")
CONFIG_FILE = str(BASE_DIR / "config.json")
BACKUP_DIR  = str(BASE_DIR / "backups")
BACKUP_MAX  = 30

DEFAULT_CONFIG = {
    "valor_diaria":       150.00,
    "backup_automatico":  True,
    "tema":               "light",
    "email_remetente":    "",
    "email_senha":        "",
    "email_smtp":         "smtp.gmail.com",
    "email_porta":        587,
    "email_destinatario": "",
    "auto_email":         False,
    "cfg_hash":           "",
    "cfg_salt":           "",
    "cfg_user":           "",
}

TIPOS_PASSAGEM   = ["Aereo","Rodoviario","Ferroviario","Maritimo","Outro"]
TIPOS_TRANSPORTE = ["Uber","99","Taxi","Onibus","Metro","VLT","Trem","Outro"]

# ── Segurança ─────────────────────────────────────────────────────────────────
def hash_senha(txt: str, salt: bytes | None = None) -> tuple[str, str]:
    if salt is None:
        salt = os.urandom(16)
    h = hashlib.scrypt(txt.strip().encode("utf-8"), salt=salt, n=2**14, r=8, p=1)
    return h.hex(), salt.hex()

def verificar_senha(txt: str, hash_hex: str, salt_hex: str) -> bool:
    h, _ = hash_senha(txt, bytes.fromhex(salt_hex))
    return h == hash_hex

def _fernet_key(senha: str, salt_hex: str) -> bytes:
    salt = bytes.fromhex(salt_hex)
    k = hashlib.scrypt(senha.strip().encode("utf-8"), salt=salt, n=2**14, r=8, p=1, dklen=32)
    return base64.urlsafe_b64encode(k)

def cifrar(texto: str, senha: str, salt_hex: str) -> str:
    if not _CRYPTO or not texto:
        return texto
    return Fernet(_fernet_key(senha, salt_hex)).encrypt(texto.encode()).decode("ascii")

def decifrar(token: str, senha: str, salt_hex: str) -> str:
    if not _CRYPTO or not token:
        return token
    try:
        return Fernet(_fernet_key(senha, salt_hex)).decrypt(token.encode("ascii")).decode()
    except Exception:
        return ""

# ── Persistência ──────────────────────────────────────────────────────────────
def load_data() -> list:
    if os.path.exists(DATA_FILE):
        try:
            with open(DATA_FILE, "r", encoding="utf-8") as f:
                dados = json.load(f)
            changed = False
            for d in dados:
                if "id" not in d:
                    d["id"] = str(uuid.uuid4())[:8]
                    changed = True
            if changed:
                save_data(dados)
            return dados
        except (json.JSONDecodeError, IOError):
            pass
    return []

def save_data(dados: list) -> bool:
    try:
        with open(DATA_FILE, "w", encoding="utf-8") as f:
            json.dump(dados, f, ensure_ascii=False, indent=4)
        return True
    except IOError:
        return False

def load_cfg() -> dict:
    if os.path.exists(CONFIG_FILE):
        try:
            with open(CONFIG_FILE, "r", encoding="utf-8") as f:
                c = json.load(f)
            for k, v in DEFAULT_CONFIG.items():
                c.setdefault(k, v)
            return c
        except (json.JSONDecodeError, IOError):
            pass
    return DEFAULT_CONFIG.copy()

def save_cfg(cfg: dict) -> bool:
    try:
        with open(CONFIG_FILE, "w", encoding="utf-8") as f:
            json.dump(cfg, f, ensure_ascii=False, indent=4)
        return True
    except IOError:
        return False

def backup_auto(cfg: dict):
    if not cfg.get("backup_automatico") or not os.path.exists(DATA_FILE):
        return
    os.makedirs(BACKUP_DIR, exist_ok=True)
    dest = os.path.join(BACKUP_DIR, f"backup_{datetime.now().strftime('%Y%m%d')}.json")
    if not os.path.exists(dest):
        try:
            shutil.copy2(DATA_FILE, dest)
        except IOError:
            pass
    try:
        arqs = sorted(Path(BACKUP_DIR).glob("backup_*.json"),
                      key=lambda p: p.stat().st_mtime, reverse=True)
        for a in arqs[BACKUP_MAX:]:
            a.unlink(missing_ok=True)
    except OSError:
        pass

# ── Lógica de negócio ─────────────────────────────────────────────────────────
def days(a: str, b: str) -> int:
    try:
        delta = (datetime.strptime(b, "%d/%m/%Y") -
                 datetime.strptime(a, "%d/%m/%Y")).days + 1
        return delta if delta >= 1 else 0
    except ValueError:
        return 0

def calc_totais(reg: dict, cfg: dict) -> dict:
    d     = days(reg.get("ida",""), reg.get("volta",""))
    vd    = cfg.get("valor_diaria", 150.0)
    val_d = d * vd
    val_p = sum(p.get("valor",0) for p in reg.get("passagens",[]))
    val_t = sum(t.get("valor",0) for t in reg.get("transportes",[]))
    val_m = sum(m.get("total",0)  for m in reg.get("materiais",[]))
    h     = reg.get("hotel",{})
    val_h = h.get("valor_noite",0)*d if isinstance(h,dict) and h.get("selecionado")=="Sim" and d>0 else 0
    return {"dias":d,"val_d":val_d,"val_p":val_p,"val_t":val_t,
            "val_m":val_m,"val_h":val_h,"total":val_d+val_p+val_t+val_m+val_h}

def salvar_registro(reg: dict, cfg: dict) -> dict:
    t = calc_totais(reg, cfg)
    reg.update({
        "id":             reg.get("id") or str(uuid.uuid4())[:8],
        "dias":           t["dias"],
        "valor":          t["val_d"],
        "val_passagens":  t["val_p"],
        "val_transporte": t["val_t"],
        "val_hotel":      t["val_h"],
        "val_materiais":  t["val_m"],
        "val_total":      t["total"],
        "data_registro":  datetime.now().strftime("%d/%m/%Y %H:%M"),
    })
    dados = load_data()
    idx = next((i for i,d in enumerate(dados) if d.get("id")==reg["id"]), None)
    if idx is not None:
        dados[idx] = reg
    else:
        dados.append(reg)
    save_data(dados)
    return reg

def totais_painel(dados: list) -> dict:
    return {
        "diarias":    sum(d.get("valor",         0) for d in dados),
        "passagens":  sum(d.get("val_passagens",  0) for d in dados),
        "transporte": sum(d.get("val_transporte", 0) for d in dados),
        "hospedagem": sum(d.get("val_hotel",      0) for d in dados),
        "materiais":  sum(d.get("val_materiais",  0) for d in dados),
        "total":      sum(d.get("val_total", d.get("valor",0)) for d in dados),
        "viagens":    len(dados),
    }

def gerar_canhoto(reg: dict, cfg: dict) -> str:
    t    = calc_totais(reg, cfg)
    vd   = cfg.get("valor_diaria", 150.0)
    h    = reg.get("hotel",{})
    nome = reg.get("nome",""); dest = reg.get("destino","")
    ida  = reg.get("ida","");  volta = reg.get("volta","")
    pix  = reg.get("pix","");  obs   = reg.get("obs","")
    part = ida   + (f" as {reg.get('hora_ida','')}"   if reg.get("hora_ida")   else "")
    ret  = volta + (f" as {reg.get('hora_volta','')}" if reg.get("hora_volta") else "")

    h_bloco = f"Utilizou Hotel  : {h.get('selecionado','Nao') if isinstance(h,dict) else 'Nao'}"
    if isinstance(h,dict) and h.get("selecionado")=="Sim":
        h_bloco += f"\nHotel Escolhido : {h.get('nome','')} (R$ {h.get('valor_noite',0):.2f}/noite)"

    p_bloco = ""
    if reg.get("passagens"):
        p_bloco = f"\nPASSAGENS ({len(reg['passagens'])} bilhete(s)):\n"
        for p in reg["passagens"]:
            dh  = p.get("data","") + (f" {p['hora']}" if p.get("hora") else "")
            loc = f" | Loc: {p['localizador']}" if p.get("localizador") else ""
            p_bloco += f"  [{p.get('tipo','')}] {p.get('trecho','')} — {dh} — R$ {p.get('valor',0):.2f}{loc}\n"

    t_bloco = ""
    if reg.get("transportes"):
        t_bloco = f"\nTRANSPORTE ({len(reg['transportes'])} lancamento(s)):\n"
        for tr in reg["transportes"]:
            t_bloco += f"  {tr.get('data','')} [{tr.get('tipo','')}] {tr.get('descricao','')} — R$ {tr.get('valor',0):.2f}\n"

    m_bloco = ""
    if reg.get("materiais"):
        m_bloco = f"\nMATERIAIS ({len(reg['materiais'])} item(ns)):\n"
        for m in reg["materiais"]:
            m_bloco += f"  {m.get('descricao','')} x{m.get('qtd',1)} — R$ {m.get('total',0):.2f}\n"

    return (
        f"\n{'='*60}\n              CANHOTO DE VIAGEM\n{'='*60}\n"
        f"Viajante        : {nome}\nDestino         : {dest}\n"
        f"Partida         : {part}\nRetorno         : {ret}\n"
        f"{h_bloco}\nDias            : {t['dias']} dia(s)\n"
        f"Valor diarias   : R$ {t['val_d']:.2f}  (R$ {vd:.2f} x {t['dias']})\n"
        f"PIX             : {pix}\n"
        + (f"Observacoes     : {obs}\n" if obs else "")
        + p_bloco + t_bloco + m_bloco
        + f"\n{'─'*60}\n"
        f"  Diarias        : R$ {t['val_d']:.2f}\n"
        f"  Passagens      : R$ {t['val_p']:.2f}\n"
        f"  Transporte     : R$ {t['val_t']:.2f}\n"
        f"  Hospedagem     : R$ {t['val_h']:.2f}\n"
        f"  Materiais      : R$ {t['val_m']:.2f}\n"
        f"  TOTAL ESTIMADO : R$ {t['total']:.2f}\n"
        f"{'─'*60}\n"
        f"Emitido em      : {datetime.now().strftime('%d/%m/%Y %H:%M')}\n"
        f"{'='*60}\n"
    )

def exportar_excel_bytes(dados: list) -> bytes | None:
    if not _EXCEL:
        return None
    import io
    wb   = Workbook()
    fill = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
    fnt  = Font(color="FFFFFF", bold=True)
    def _hdr(ws, cols):
        ws.append(cols)
        for c in ws[1]:
            c.fill=fill; c.font=fnt; c.alignment=Alignment(horizontal="center")

    ws1 = wb.active; ws1.title = "Viagens"
    _hdr(ws1,["Nome","Destino","Ida","Volta","Hotel","Dias",
              "Diarias","Passagens","Transporte","Hospedagem","Materiais","TOTAL","PIX"])
    for d in dados:
        h  = d.get("hotel",{})
        vt = d.get("val_total", d.get("valor",0))
        ws1.append([d.get("nome",""), d.get("destino",""), d.get("ida",""), d.get("volta",""),
                    h.get("nome","") if isinstance(h,dict) else "",
                    d.get("dias",0), d.get("valor",0), d.get("val_passagens",0),
                    d.get("val_transporte",0), d.get("val_hotel",0),
                    d.get("val_materiais",0), vt, d.get("pix","")])

    ws2 = wb.create_sheet("Passagens")
    _hdr(ws2,["Viajante","Destino","Tipo","Trecho","Data","Valor","Localizador"])
    for d in dados:
        for p in d.get("passagens",[]):
            ws2.append([d.get("nome",""), d.get("destino",""), p.get("tipo",""),
                        p.get("trecho",""), p.get("data",""),
                        p.get("valor",0), p.get("localizador","")])

    ws3 = wb.create_sheet("Transporte")
    _hdr(ws3,["Viajante","Destino","Data","Tipo","Descricao","Valor"])
    for d in dados:
        for t in d.get("transportes",[]):
            ws3.append([d.get("nome",""), d.get("destino",""), t.get("data",""),
                        t.get("tipo",""), t.get("descricao",""), t.get("valor",0)])

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()

def enviar_email(cfg: dict, assunto: str, corpo: str, email_senha_plain: str="") -> tuple[bool,str]:
    rem  = cfg.get("email_remetente","").strip()
    pwd  = email_senha_plain or cfg.get("email_senha","").strip()
    dest = cfg.get("email_destinatario","").strip()
    srv  = cfg.get("email_smtp","smtp.gmail.com")
    try:
        porta = int(cfg.get("email_porta",587))
    except (ValueError,TypeError):
        porta = 587
    if not all([rem,pwd,dest]):
        return False, "Configure remetente, senha e destinatario."
    try:
        msg = MIMEMultipart("alternative")
        msg["Subject"] = assunto; msg["From"] = rem; msg["To"] = dest
        now = datetime.now().strftime("%d/%m/%Y %H:%M")
        msg.attach(MIMEText(f"Canhoto gerado em {now}.\n\n{corpo}", "plain","utf-8"))
        msg.attach(MIMEText(
            f"<html><body><p><b>{html.escape(assunto)}</b> — {html.escape(now)}</p>"
            f"<pre style='font-family:Courier'>{html.escape(corpo)}</pre></body></html>",
            "html","utf-8"))
        with smtplib.SMTP(srv, porta, timeout=15) as s:
            s.ehlo(); s.starttls(); s.ehlo(); s.login(rem, pwd)
            s.sendmail(rem, dest, msg.as_string())
        return True, f"E-mail enviado para {dest}"
    except smtplib.SMTPAuthenticationError:
        return False, "Senha incorreta. Use Senha de App para Gmail."
    except Exception as e:
        return False, str(e)

def fmt_brl(v: float) -> str:
    return f"R$ {v:,.2f}".replace(",","X").replace(".",",").replace("X",".")
